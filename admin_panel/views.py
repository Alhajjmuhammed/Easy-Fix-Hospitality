from django.shortcuts import render, redirect, get_object_or_404
from .restaurant_utils import get_restaurant_context, get_current_restaurant, filter_data_by_restaurant
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required


def _safe_csv(value):
    """Prevent CSV formula injection: prefix dangerous leading characters."""
    s = str(value) if value is not None else ''
    if s and s[0] in ('=', '+', '-', '@', '|', '%'):
        return '\t' + s
    return s
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.db import models, transaction
from django.utils import timezone
from datetime import timedelta
from django.core.paginator import Paginator
from accounts.models import User, Role, get_owner_filter
from restaurant.models import Product, MainCategory, SubCategory, TableInfo, validate_image_file
from restaurant.models_restaurant import Restaurant
from orders.models import Order, OrderItem
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_http_methods
from django.template.loader import render_to_string
from django.core.exceptions import PermissionDenied, ValidationError
from django.db.models import ProtectedError
import json
import qrcode
import io
import csv
from decimal import Decimal, InvalidOperation
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import tempfile
import os
try:
    import openpyxl
except ImportError:
    openpyxl = None

import logging
logger = logging.getLogger(__name__)


def _cleanup_order_dependencies(order_ids):
    """Delete/nullify all PROTECT FK references blocking order deletion on PostgreSQL."""
    if not order_ids:
        return
    try:
        from cashier.models import Payment, OrderItemPayment, VoidTransaction
        payment_ids = list(Payment.objects.filter(order_id__in=order_ids).values_list('id', flat=True))
        if payment_ids:
            VoidTransaction.objects.filter(original_payment_id__in=payment_ids).delete()
            OrderItemPayment.objects.filter(payment_id__in=payment_ids).delete()
            Payment.objects.filter(id__in=payment_ids).delete()
    except ImportError:
        pass
    try:
        from waste_management.models import FoodWasteLog, OrderCostBreakdown
        FoodWasteLog.objects.filter(order_id__in=order_ids).update(order=None)
        OrderCostBreakdown.objects.filter(order_id__in=order_ids).delete()
    except ImportError:
        pass


def _cleanup_product_dependencies(product_ids):
    """Nullify/delete all PROTECT FK references blocking product deletion on PostgreSQL."""
    if not product_ids:
        return
    OrderItem.objects.filter(product_id__in=product_ids).update(product=None)
    try:
        from waste_management.models import FoodWasteLog, ProductCostSettings
        FoodWasteLog.objects.filter(product_id__in=product_ids).update(product=None)
        ProductCostSettings.objects.filter(product_id__in=product_ids).delete()
    except ImportError:
        pass
    try:
        from reports.models import ProductSalesDetail
        ProductSalesDetail.objects.filter(product_id__in=product_ids).update(product=None)
    except ImportError:
        pass


def get_production_qr_url(request, qr_code):
    """Build absolute QR URL using the current request scheme and host.
    SECURE_PROXY_SSL_HEADER in production_settings ensures HTTPS is detected correctly."""
    return request.build_absolute_uri(f'/r/{qr_code}/')


@login_required
def admin_dashboard(request):
    """Main admin dashboard view - accessible by administrators, owners, and managers"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator, Owner, or Manager privileges required.")
        return redirect('restaurant:home')

    try:
        # Get current restaurant context
        current_restaurant = None
        view_all_restaurants = request.session.get('view_all_restaurants', False)
        selected_restaurant_id = request.session.get('selected_restaurant_id')
        
        if selected_restaurant_id and not view_all_restaurants:
            # selected_restaurant_id stores User (owner) ID, not Restaurant ID
            try:
                selected_user = User.objects.get(id=selected_restaurant_id)
                
                # Get the Restaurant object for this user
                if selected_user.is_branch_owner():
                    current_restaurant = Restaurant.objects.filter(branch_owner=selected_user, is_main_restaurant=False).first()
                else:
                    current_restaurant = Restaurant.objects.filter(main_owner=selected_user, is_main_restaurant=True).first()
                
                # Verify user can access this restaurant
                if current_restaurant and not current_restaurant.can_user_access(request.user):
                    current_restaurant = None
                    if 'selected_restaurant_id' in request.session:
                        del request.session['selected_restaurant_id']
            except User.DoesNotExist:
                current_restaurant = None
                if 'selected_restaurant_id' in request.session:
                    del request.session['selected_restaurant_id']
        
        # If no current restaurant selected, auto-select for non-main owners
        if not current_restaurant and not view_all_restaurants:
            if request.user.is_branch_owner():
                # Branch owners should see their restaurant by default
                user_restaurants = Restaurant.objects.filter(branch_owner=request.user)
                if user_restaurants.exists():
                    current_restaurant = user_restaurants.first()
                    request.session['selected_restaurant_id'] = request.user.id  # Store User ID (convention)
            elif request.user.is_owner() and not request.user.is_main_owner():
                # Legacy owners - try to find their restaurant via TableInfo
                owner_filter = get_owner_filter(request.user)
                if owner_filter:
                    # Try to find restaurant associated with this owner
                    user_restaurants = Restaurant.objects.filter(
                        Q(main_owner=request.user) | Q(branch_owner=request.user)
                    )
                    if user_restaurants.exists():
                        current_restaurant = user_restaurants.first()
                        request.session['selected_restaurant_id'] = request.user.id  # Store User ID (convention)
        
        # Calculate statistics based on context
        if request.user.is_administrator():
            # Administrators see everything
            if view_all_restaurants or not current_restaurant:
                # All restaurants view
                total_users = User.objects.count()
                total_orders = Order.objects.count()
                total_products = Product.objects.count()
                total_tables = TableInfo.objects.count()
                context_name = "All Restaurants (Admin)"
            else:
                # Specific restaurant for admin
                # Count: owner + their staff
                if current_restaurant.is_main_restaurant:
                    staff_count = User.objects.filter(owner=current_restaurant.main_owner).count()
                    total_users = 1 + staff_count  # main owner + staff
                else:
                    staff_count = User.objects.filter(owner=current_restaurant.branch_owner).count()
                    total_users = 1 + staff_count  # branch owner + staff
                total_orders = Order.objects.filter(
                    Q(table_info__restaurant=current_restaurant) |
                    Q(table_info__owner__in=[current_restaurant.main_owner, current_restaurant.branch_owner])
                ).count()
                total_products = Product.objects.filter(
                    Q(main_category__restaurant=current_restaurant) |
                    Q(main_category__owner__in=[current_restaurant.main_owner, current_restaurant.branch_owner])
                ).count()
                total_tables = TableInfo.objects.filter(
                    Q(restaurant=current_restaurant) |
                    Q(owner__in=[current_restaurant.main_owner, current_restaurant.branch_owner])
                ).count()
                context_name = f"{current_restaurant.name} (Admin)"
                
        elif request.user.is_main_owner():
            # Main owners - can see all or individual restaurants
            if view_all_restaurants or not current_restaurant:
                # All restaurants owned by this main owner
                owned_restaurants = Restaurant.objects.filter(main_owner=request.user)
                # Pre-compute branch owner IDs at DB level — avoids N+1 FK access per restaurant
                branch_owner_ids = list(
                    owned_restaurants.filter(is_main_restaurant=False)
                    .values_list('branch_owner_id', flat=True)
                )
                all_owner_ids = list(
                    owned_restaurants.exclude(branch_owner__isnull=True)
                    .values_list('branch_owner_id', flat=True)
                )
                # Count: main owner (1) + main staff + branch owners + branch staff
                main_staff = User.objects.filter(owner=request.user).count()
                branch_count = len(branch_owner_ids)
                branch_staff = User.objects.filter(owner__in=branch_owner_ids).count()
                total_users = 1 + main_staff + branch_count + branch_staff
                total_orders = Order.objects.filter(
                    Q(table_info__restaurant__in=owned_restaurants) |
                    Q(table_info__owner__in=all_owner_ids)
                ).count()
                total_products = Product.objects.filter(
                    Q(main_category__restaurant__in=owned_restaurants) |
                    Q(main_category__owner__in=all_owner_ids)
                ).count()
                total_tables = TableInfo.objects.filter(
                    Q(restaurant__in=owned_restaurants) |
                    Q(owner__in=all_owner_ids)
                ).count()
                context_name = f"All Restaurants ({branch_count + 1} locations)"
            else:
                # Specific restaurant
                if current_restaurant.main_owner != request.user:
                    raise PermissionDenied("You don't have access to this restaurant.")
                
                # Count branch owner + their staff
                if current_restaurant.is_main_restaurant:
                    # Main restaurant: count main owner + their staff
                    total_users = 1 + User.objects.filter(owner=current_restaurant.main_owner).count()
                else:
                    # Branch: count branch owner + their staff
                    total_users = 1 + User.objects.filter(owner=current_restaurant.branch_owner).count()
                total_orders = Order.objects.filter(
                    Q(table_info__restaurant=current_restaurant) |
                    Q(table_info__owner=current_restaurant.branch_owner)
                ).count()
                total_products = Product.objects.filter(
                    Q(main_category__restaurant=current_restaurant) |
                    Q(main_category__owner=current_restaurant.branch_owner)
                ).count()
                total_tables = TableInfo.objects.filter(
                    Q(restaurant=current_restaurant) |
                    Q(owner=current_restaurant.branch_owner)
                ).count()
                context_name = current_restaurant.name
                
        elif request.user.is_manager():
            # Managers - see their restaurant owner's data (same as owner sees)
            if current_restaurant:
                if not current_restaurant.can_user_access(request.user):
                    raise PermissionDenied("You don't have access to this restaurant.")
                
                # Get the owner this manager belongs to
                manager_owner = request.user.owner
                
                # Count: owner (1) + all staff under that owner (including this manager)
                staff_count = User.objects.filter(owner=manager_owner).count()
                total_users = 1 + staff_count
                
                total_orders = Order.objects.filter(
                    Q(table_info__restaurant=current_restaurant) |
                    Q(table_info__owner=manager_owner)
                ).count()
                total_products = Product.objects.filter(
                    Q(main_category__restaurant=current_restaurant) |
                    Q(main_category__owner=manager_owner)
                ).count()
                total_tables = TableInfo.objects.filter(
                    Q(restaurant=current_restaurant) |
                    Q(owner=manager_owner)
                ).count()
                context_name = current_restaurant.name
            else:
                # Fallback - use manager's owner
                manager_owner = request.user.owner
                if manager_owner:
                    total_users = User.objects.filter(owner=manager_owner).count() + 1
                    total_orders = Order.objects.filter(
                        Q(table_info__owner=manager_owner) |
                        Q(table_info__restaurant__main_owner=manager_owner) |
                        Q(table_info__restaurant__branch_owner=manager_owner)
                    ).distinct().count()
                    total_products = Product.objects.filter(
                        Q(main_category__owner=manager_owner) |
                        Q(main_category__restaurant__main_owner=manager_owner) |
                        Q(main_category__restaurant__branch_owner=manager_owner)
                    ).distinct().count()
                    total_tables = TableInfo.objects.filter(
                        Q(owner=manager_owner) |
                        Q(restaurant__main_owner=manager_owner) |
                        Q(restaurant__branch_owner=manager_owner)
                    ).distinct().count()
                    context_name = manager_owner.get_restaurant_name()
                else:
                    raise PermissionDenied("No restaurant access found.")
                    
        else:
            # Branch owners and legacy owners
            if current_restaurant:
                if not current_restaurant.can_user_access(request.user):
                    raise PermissionDenied("You don't have access to this restaurant.")
                
                # Count: owner themselves (1) + their staff
                staff_count = User.objects.filter(owner=request.user).count()
                total_users = 1 + staff_count
                total_orders = Order.objects.filter(
                    Q(table_info__restaurant=current_restaurant) |
                    Q(table_info__owner=request.user)
                ).count()
                total_products = Product.objects.filter(
                    Q(main_category__restaurant=current_restaurant) |
                    Q(main_category__owner=request.user)
                ).count()
                total_tables = TableInfo.objects.filter(
                    Q(restaurant=current_restaurant) |
                    Q(owner=request.user)
                ).count()
                context_name = current_restaurant.name
            else:
                # Fallback to legacy owner filter
                owner_filter = get_owner_filter(request.user)
                if owner_filter:
                    total_users = User.objects.filter(owner=owner_filter).count() + 1
                    total_orders = Order.objects.filter(
                        Q(table_info__owner=owner_filter) |
                        Q(table_info__restaurant__main_owner=owner_filter) |
                        Q(table_info__restaurant__branch_owner=owner_filter)
                    ).distinct().count()
                    total_products = Product.objects.filter(
                        Q(main_category__owner=owner_filter) |
                        Q(main_category__restaurant__main_owner=owner_filter) |
                        Q(main_category__restaurant__branch_owner=owner_filter)
                    ).distinct().count()
                    total_tables = TableInfo.objects.filter(
                        Q(owner=owner_filter) |
                        Q(restaurant__main_owner=owner_filter) |
                        Q(restaurant__branch_owner=owner_filter)
                    ).distinct().count()
                    context_name = request.user.get_restaurant_name()
                else:
                    raise PermissionDenied("No restaurant access found.")
        
        # Calculate time-based statistics
        seven_days_ago = timezone.now() - timedelta(days=7)
        today = timezone.now().date()
        
        if current_restaurant and not view_all_restaurants:
            # Single restaurant stats — 1 aggregate instead of 3 queries
            _restaurant_q = (
                Q(table_info__restaurant=current_restaurant) |
                Q(table_info__owner__in=[current_restaurant.main_owner, current_restaurant.branch_owner])
            )
            _stats = Order.objects.filter(_restaurant_q).aggregate(
                recent_orders=Count('id', filter=Q(created_at__gte=seven_days_ago)),
                today_revenue=Sum('total_amount', filter=Q(created_at__date=today)),
                pending_orders=Count('id', filter=Q(status='pending')),
            )
            recent_orders = _stats['recent_orders'] or 0
            today_revenue = _stats['today_revenue'] or 0
            pending_orders = _stats['pending_orders'] or 0
        elif request.user.is_main_owner() and view_all_restaurants:
            # All restaurants for main owner — 1 aggregate instead of 3 queries
            owned_restaurants = Restaurant.objects.filter(main_owner=request.user)
            _branch_owners = list(owned_restaurants.values_list('branch_owner', flat=True))
            _owner_q = (
                Q(table_info__restaurant__in=owned_restaurants) |
                Q(table_info__owner__in=_branch_owners)
            )
            _stats = Order.objects.filter(_owner_q).aggregate(
                recent_orders=Count('id', filter=Q(created_at__gte=seven_days_ago)),
                today_revenue=Sum('total_amount', filter=Q(created_at__date=today)),
                pending_orders=Count('id', filter=Q(status='pending')),
            )
            recent_orders = _stats['recent_orders'] or 0
            today_revenue = _stats['today_revenue'] or 0
            pending_orders = _stats['pending_orders'] or 0
        else:
            # Fallback - use owner filter or all
            if not request.user.is_administrator():
                owner_filter = get_owner_filter(request.user)
                # 1 aggregate instead of 3 queries
                _oq_fb = (
                    Q(table_info__owner=owner_filter) |
                    Q(table_info__restaurant__main_owner=owner_filter) |
                    Q(table_info__restaurant__branch_owner=owner_filter)
                )
                _stats = Order.objects.filter(_oq_fb).aggregate(
                    recent_orders=Count('id', filter=Q(created_at__gte=seven_days_ago)),
                    today_revenue=Sum('total_amount', filter=Q(created_at__date=today)),
                    pending_orders=Count('id', filter=Q(status='pending')),
                )
                recent_orders = _stats['recent_orders'] or 0
                today_revenue = _stats['today_revenue'] or 0
                pending_orders = _stats['pending_orders'] or 0
            else:
                # Administrator - all data — 1 aggregate instead of 3 queries
                _stats = Order.objects.aggregate(
                    recent_orders=Count('id', filter=Q(created_at__gte=seven_days_ago)),
                    today_revenue=Sum('total_amount', filter=Q(created_at__date=today)),
                    pending_orders=Count('id', filter=Q(status='pending')),
                )
                recent_orders = _stats['recent_orders'] or 0
                today_revenue = _stats['today_revenue'] or 0
                pending_orders = _stats['pending_orders'] or 0
            
    except PermissionDenied as e:
        logger.error(f'Permission denied in dashboard: {str(e)}')
        messages.error(request, 'Access denied. Please contact an administrator.')
        return redirect('restaurant:home')
    except Exception as e:
        logger.error(f'Error loading dashboard: {str(e)}')
        messages.error(request, 'Error loading dashboard. Please try again.')
        return redirect('restaurant:home')

    # Get restaurant context using utility
    session_restaurant_id = request.session.get('selected_restaurant_id')
    restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
    current_restaurant = restaurant_context['current_restaurant']
    view_all_restaurants = restaurant_context['view_all_restaurants']
    
    # Restaurant name for display
    context_name = restaurant_context['restaurant_name']

    # Get current restaurant's QR code
    current_qr_code = None
    if current_restaurant and current_restaurant.qr_code:
        current_qr_code = current_restaurant.qr_code
    elif request.user.restaurant_qr_code:
        # Fallback to user's QR code for legacy support
        current_qr_code = request.user.restaurant_qr_code
    elif request.user.is_manager() and request.user.owner and request.user.owner.restaurant_qr_code:
        # Managers use their owner's QR code
        current_qr_code = request.user.owner.restaurant_qr_code
    
    # Combine dashboard data with restaurant context
    context = {
        'total_users': total_users,
        'total_orders': total_orders,
        'total_products': total_products,
        'total_tables': total_tables,
        'recent_orders': recent_orders,
        'today_revenue': today_revenue,
        'pending_orders': pending_orders,
        'current_qr_code': current_qr_code,  # Current restaurant's QR code
        **restaurant_context,  # Include all restaurant context variables
    }

    return render(request, 'admin_panel/dashboard.html', context)


@login_required
def manage_users(request):
    """User management view"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator/Owner privileges required.")
        return redirect('restaurant:home')

    try:
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        view_all_restaurants = restaurant_context['view_all_restaurants']
        
        if current_restaurant and not view_all_restaurants:
            # Filter users related to specific selected restaurant
            if current_restaurant.is_main_restaurant:
                # Main restaurant: show main owner, branch owner, and staff under main owner
                users = User.objects.filter(
                    Q(owned_restaurants=current_restaurant) |  # Main owners
                    Q(managed_restaurant=current_restaurant) |  # Branch owners of this main restaurant
                    Q(owner=current_restaurant.main_owner)  # Staff under main owner
                ).exclude(id=request.user.id).distinct().order_by('-date_joined')  # Exclude current user
            else:
                # Branch restaurant: only show branch owner and staff under branch owner
                users = User.objects.filter(
                    Q(managed_restaurant=current_restaurant) |  # Branch owner of this branch
                    Q(owner=current_restaurant.branch_owner)  # Staff under branch owner
                ).exclude(id=request.user.id).distinct().order_by('-date_joined')  # Exclude current user
            
            # Managers can only see staff, not owners or peer managers
            if request.user.is_manager():
                users = users.exclude(
                    Q(role__name__in=['administrator', 'main_owner', 'branch_owner', 'owner', 'manager']) |
                    Q(is_superuser=True)
                )
            roles = Role.objects.exclude(name='administrator')
            
            # Filter roles based on user type and subscription plan
            if request.user.is_branch_owner():
                # Branch owners can ONLY create staff roles
                roles = roles.exclude(name__in=['main_owner', 'branch_owner', 'owner'])
            elif request.user.is_main_owner():
                # Main owners cannot create another main_owner, hide legacy 'owner' role
                roles = roles.exclude(name__in=['main_owner', 'owner'])
                # On SINGLE plan, also hide branch_owner
                if current_restaurant and current_restaurant.subscription_plan == 'SINGLE':
                    roles = roles.exclude(name='branch_owner')
            elif request.user.is_manager():
                # Managers can ONLY create regular staff roles (not owners, managers, or customers)
                roles = roles.exclude(name__in=['administrator', 'main_owner', 'branch_owner', 'owner', 'manager', 'customer'])
            elif current_restaurant and current_restaurant.subscription_plan == 'SINGLE':
                # SINGLE plan: Only allow staff roles (no owner roles)
                roles = roles.exclude(name__in=['main_owner', 'branch_owner', 'owner'])
                
        elif request.user.is_administrator():
            # Administrators see all users (except themselves)
            users = User.objects.exclude(id=request.user.id).order_by('-date_joined')
            roles = Role.objects.all()
        else:
            # Get users from all accessible restaurants
            accessible_restaurants = restaurant_context['accessible_restaurants']
            
            if accessible_restaurants.exists():
                user_query = Q()
                for restaurant in accessible_restaurants:
                    user_query |= (
                        Q(owned_restaurants=restaurant) |  # Main owners
                        Q(managed_restaurant=restaurant) |  # Branch owners
                        Q(owner=restaurant.main_owner) |  # Staff under main owner
                        Q(owner=restaurant.branch_owner)  # Staff under branch owner
                    )
                
                users = User.objects.filter(user_query).exclude(id=request.user.id).distinct().order_by('-date_joined')  # Exclude current user
                
                # Managers can only see staff, not owners or peer managers
                if request.user.is_manager():
                    users = users.exclude(
                        Q(role__name__in=['administrator', 'main_owner', 'branch_owner', 'owner', 'manager']) |
                        Q(is_superuser=True)
                    )
                
                roles = Role.objects.exclude(name='administrator')
                
                # Filter roles based on user type and subscription plan
                if request.user.is_branch_owner():
                    # Branch owners can ONLY create staff roles
                    roles = roles.exclude(name__in=['main_owner', 'branch_owner', 'owner'])
                elif request.user.is_main_owner():
                    # Main owners cannot create another main_owner, hide legacy 'owner' role
                    roles = roles.exclude(name__in=['main_owner', 'owner'])
                    # On SINGLE plan, also hide branch_owner
                    if current_restaurant and current_restaurant.subscription_plan == 'SINGLE':
                        roles = roles.exclude(name='branch_owner')
                elif request.user.is_manager():
                    # Managers can ONLY create regular staff roles (not owners, managers, or customers)
                    roles = roles.exclude(name__in=['administrator', 'main_owner', 'branch_owner', 'owner', 'manager', 'customer'])
                elif accessible_restaurants.filter(subscription_plan='SINGLE').exists():
                    # If viewing SINGLE plan restaurant, exclude owner roles
                    if current_restaurant and current_restaurant.subscription_plan == 'SINGLE':
                        roles = roles.exclude(name__in=['main_owner', 'branch_owner', 'owner'])
            else:
                # No accessible restaurants for this non-admin user — show nothing (defensive fallback)
                users = User.objects.none()
                roles = Role.objects.none()
            
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    # Apply filters
    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '').strip()
    status_filter = request.GET.get('status', '').strip()
    restaurant_filter = request.GET.get('restaurant', '').strip()

    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    if role_filter:
        # Filter by role using the role relationship
        if role_filter == 'administrator':
            users = users.filter(role__name='administrator')
        elif role_filter == 'owner':
            # Include all owner types: main_owner, branch_owner, and legacy owner
            users = users.filter(role__name__in=['owner', 'main_owner', 'branch_owner'])
        elif role_filter == 'main_owner':
            users = users.filter(role__name='main_owner')
        elif role_filter == 'branch_owner':
            users = users.filter(role__name='branch_owner')
        elif role_filter == 'kitchen':
            users = users.filter(role__name='kitchen')
        elif role_filter == 'bar':
            users = users.filter(role__name='bar')
        elif role_filter == 'customer_care':
            users = users.filter(role__name='customer_care')
        elif role_filter == 'cashier':
            users = users.filter(role__name='cashier')
        elif role_filter == 'customer':
            users = users.filter(role__name='customer')

    if status_filter:
        if status_filter == 'active':
            users = users.filter(is_active=True)
        elif status_filter == 'inactive':
            users = users.filter(is_active=False)

    if restaurant_filter:
        from restaurant.models_restaurant import Restaurant
        try:
            filter_restaurant = Restaurant.objects.get(id=restaurant_filter)
            if filter_restaurant.is_main_restaurant:
                users = users.filter(
                    Q(owned_restaurants=filter_restaurant) |
                    Q(managed_restaurant=filter_restaurant) |
                    Q(owner=filter_restaurant.main_owner)
                ).distinct()
            else:
                users = users.filter(
                    Q(managed_restaurant=filter_restaurant) |
                    Q(owner=filter_restaurant.branch_owner)
                ).distinct()
        except Restaurant.DoesNotExist:
            pass

    # Pagination
    try:
        per_page = int(request.GET.get('per_page', 20))
    except (ValueError, TypeError):
        per_page = 20
    paginator = Paginator(users, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Annotate each role with user count to avoid N+1 in template
    roles = roles.annotate(user_count=Count('user'))

    context = {
        'users': page_obj.object_list,
        'page_obj': page_obj,
        'roles': roles,
        'is_owner_access': (request.user.is_owner() or request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()) and not request.user.is_administrator(),
        **restaurant_context,  # Include restaurant context
    }

    return render(request, 'admin_panel/manage_users.html', context)


@login_required
def manage_products(request):
    """Product management view - accessible by administrators and owners"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('restaurant:home')

    # Initialize variables
    main_categories = MainCategory.objects.none()
    all_products = Product.objects.none()
    base_categories = MainCategory.objects.none()

    try:
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        view_all_restaurants = restaurant_context['view_all_restaurants']
        
        # Get filter parameters first
        search_query = request.GET.get('search', '').strip()
        category_filter = request.GET.get('category', '').strip()
        restaurant_filter = request.GET.get('restaurant', '').strip()
        try:
            per_page = int(request.GET.get('per_page', 10))
        except (ValueError, TypeError):
            per_page = 10

        # Determine which restaurant to filter by
        target_restaurant = None
        if restaurant_filter:
            try:
                from restaurant.models_restaurant import Restaurant
                target_restaurant = Restaurant.objects.get(id=restaurant_filter)
                # Security: verify user has access to this restaurant
                if not request.user.is_administrator():
                    if not restaurant_context['accessible_restaurants'].filter(id=target_restaurant.id).exists():
                        target_restaurant = None
            except Restaurant.DoesNotExist:
                target_restaurant = None
        elif current_restaurant and not view_all_restaurants:
            target_restaurant = current_restaurant

        if target_restaurant:
            # Filter by specific restaurant (either from filter or session)
            if target_restaurant.is_main_restaurant:
                # Main restaurant: show categories/products assigned to it OR owned by main owner with no restaurant
                base_categories = MainCategory.objects.filter(
                    is_active=True
                ).filter(
                    Q(restaurant=target_restaurant) | 
                    Q(owner=target_restaurant.main_owner, restaurant__isnull=True)
                ).order_by('name')
            else:
                # Branch restaurant: only show categories/products specifically assigned to this branch
                # OR created by the branch owner (not main owner)
                branch_query = Q(restaurant=target_restaurant)
                if target_restaurant.branch_owner:
                    branch_query |= Q(owner=target_restaurant.branch_owner, restaurant__isnull=True)
                    
                base_categories = MainCategory.objects.filter(
                    is_active=True
                ).filter(branch_query).order_by('name')
        elif request.user.is_administrator():
            # Administrator sees all products and categories
            base_categories = MainCategory.objects.filter(is_active=True).order_by('name')
        else:
            # Get user's accessible restaurants (view all mode)
            accessible_restaurants = restaurant_context['accessible_restaurants']
            
            if accessible_restaurants.exists():
                # Build query for accessible restaurants and legacy owner field
                restaurant_query = Q()
                owner_query = Q()
                
                for restaurant in accessible_restaurants:
                    restaurant_query |= Q(restaurant=restaurant)
                    if restaurant.main_owner:
                        owner_query |= Q(owner=restaurant.main_owner, restaurant__isnull=True)
                    if restaurant.branch_owner:
                        owner_query |= Q(owner=restaurant.branch_owner, restaurant__isnull=True)
                
                base_categories = MainCategory.objects.filter(
                    is_active=True
                ).filter(restaurant_query | owner_query).order_by('name')
            else:
                base_categories = MainCategory.objects.none()
        
        # Start with base categories
        main_categories = base_categories

        # Apply category filter — scope directly to the existing tenant-scoped queryset
        if category_filter:
            main_categories = main_categories.filter(id=category_filter)

        # Get all products for the filtered categories
        all_products = Product.objects.filter(
            main_category__in=main_categories
        ).select_related('main_category', 'sub_category').order_by('name')

        # Apply search filter
        if search_query:
            all_products = all_products.filter(name__icontains=search_query)
        
        # Pre-compute product counts per category — single GROUP BY, avoids 2 queries × N categories
        # order_by() clears the default ordering to prevent Django from adding 'name' into GROUP BY
        _product_counts = {
            row['main_category_id']: row['count']
            for row in all_products.order_by().values('main_category_id').annotate(count=Count('id'))
        }

        # Add pagination for each category
        paginated_categories = []
        for category in main_categories:
            total_count = _product_counts.get(category.id, 0)
            if total_count == 0:
                continue
            category_products = all_products.filter(main_category=category)

            # Get page number for this category (default 1)
            page_param = f'page_{category.id}'
            page_number = request.GET.get(page_param, 1)

            # Paginate products
            paginator = Paginator(category_products, per_page)
            page_obj = paginator.get_page(page_number)

            paginated_categories.append({
                'category': category,
                'products': page_obj,
                'page_param': page_param,
                'total_count': total_count,
            })
            
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    context = {
        'main_categories': paginated_categories,
        'all_main_categories': base_categories,  # Use base_categories for the dropdown (all accessible categories)
        **restaurant_context,  # Include restaurant context
    }

    return render(request, 'admin_panel/manage_products.html', context)


@login_required
def manage_orders(request):
    """Order management view with status-based tabs - accessible by administrators and owners"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('restaurant:home')

    try:
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        view_all_restaurants = restaurant_context['view_all_restaurants']
        
        # Get filter parameters
        search_query = request.GET.get('search', '').strip()
        restaurant_filter = request.GET.get('restaurant', '').strip()
        payment_status_filter = request.GET.get('payment_status', '').strip()
        date_from = request.GET.get('date_from', '').strip()
        date_to = request.GET.get('date_to', '').strip()
        period = request.GET.get('period', 'all')

        # Determine which restaurant to filter by
        target_restaurant = None
        if restaurant_filter:
            try:
                from restaurant.models_restaurant import Restaurant
                target_restaurant = Restaurant.objects.get(id=restaurant_filter)
                # Security: verify user has access to this restaurant
                if not request.user.is_administrator():
                    if not restaurant_context['accessible_restaurants'].filter(id=target_restaurant.id).exists():
                        target_restaurant = None
            except Restaurant.DoesNotExist:
                target_restaurant = None
        elif current_restaurant and not view_all_restaurants:
            target_restaurant = current_restaurant
        
        if target_restaurant:
            # Filter orders by specific selected restaurant
            if target_restaurant.is_main_restaurant:
                base_orders = Order.objects.filter(
                    Q(table_info__restaurant=target_restaurant) |
                    Q(table_info__owner=target_restaurant.main_owner, table_info__restaurant__isnull=True)
                )
            else:
                base_orders = Order.objects.filter(
                    Q(table_info__restaurant=target_restaurant) |
                    Q(table_info__owner=target_restaurant.branch_owner, table_info__restaurant__isnull=True)
                )
        elif request.user.is_administrator():
            base_orders = Order.objects.all()
        else:
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if accessible_restaurants.exists():
                order_query = Q()
                for restaurant in accessible_restaurants:
                    order_query |= (
                        Q(table_info__restaurant=restaurant) |
                        Q(table_info__owner=restaurant.main_owner, table_info__restaurant__isnull=True) |
                        Q(table_info__owner=restaurant.branch_owner, table_info__restaurant__isnull=True)
                    )
                base_orders = Order.objects.filter(order_query)
            else:
                base_orders = Order.objects.none()
        
        # Eager-load related objects once — prevents N+1 queries in the template
        base_orders = base_orders.select_related(
            'table_info', 'table_info__owner', 'ordered_by', 'confirmed_by'
        ).prefetch_related('order_items__product')

        # Apply search filter
        if search_query:
            base_orders = base_orders.filter(
                Q(order_number__icontains=search_query) |
                Q(table_info__tbl_no__icontains=search_query) |
                Q(ordered_by__username__icontains=search_query) |
                Q(ordered_by__first_name__icontains=search_query) |
                Q(ordered_by__last_name__icontains=search_query)
            )
        
        # Apply payment status filter
        if payment_status_filter:
            base_orders = base_orders.filter(payment_status=payment_status_filter)

        # Apply period filter (takes priority over manual dates when set)
        from django.utils import timezone as _tz
        _today = _tz.now().date()
        if period == 'today':
            base_orders = base_orders.filter(created_at__date=_today)
        elif period == 'week':
            from datetime import timedelta as _td
            _week_start = _today - _td(days=_today.weekday())
            base_orders = base_orders.filter(created_at__date__gte=_week_start)
        elif period == 'month':
            _month_start = _today.replace(day=1)
            base_orders = base_orders.filter(created_at__date__gte=_month_start)
        elif period == 'year':
            _year_start = _today.replace(month=1, day=1)
            base_orders = base_orders.filter(created_at__date__gte=_year_start)
        else:
            # 'all' — apply manual date range if provided
            if date_from:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                base_orders = base_orders.filter(created_at__gte=date_from_obj)
            if date_to:
                from datetime import datetime, timedelta
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                date_to_obj = date_to_obj + timedelta(days=1)
                base_orders = base_orders.filter(created_at__lt=date_to_obj)

    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    # Organize orders by status
    pending_orders   = base_orders.filter(status='pending').order_by('-created_at')
    confirmed_orders = base_orders.filter(status='confirmed').order_by('-created_at')
    preparing_orders = base_orders.filter(status='preparing').order_by('-created_at')
    ready_orders     = base_orders.filter(status='ready').order_by('-created_at')
    served_orders    = base_orders.filter(status='served').order_by('-created_at')
    cancelled_orders = base_orders.filter(status='cancelled').order_by('-created_at')
    
    # Counts — single GROUP BY query
    _counts_qs = base_orders.values('status').annotate(count=Count('id'))
    _counts = {row['status']: row['count'] for row in _counts_qs}
    pending_count   = _counts.get('pending', 0)
    confirmed_count = _counts.get('confirmed', 0)
    preparing_count = _counts.get('preparing', 0)
    ready_count     = _counts.get('ready', 0)
    served_count    = _counts.get('served', 0)
    cancelled_count = _counts.get('cancelled', 0)
    total_count     = sum(_counts.values())

    # Paginate each tab separately
    from django.core.paginator import Paginator
    try:
        _per_page = int(request.GET.get('per_page', 25))
        if _per_page not in [10, 25, 50, 100]:
            _per_page = 25
    except (ValueError, TypeError):
        _per_page = 25
    pending_page   = Paginator(pending_orders,   _per_page).get_page(request.GET.get('pending_page',   1))
    confirmed_page = Paginator(confirmed_orders, _per_page).get_page(request.GET.get('confirmed_page', 1))
    preparing_page = Paginator(preparing_orders, _per_page).get_page(request.GET.get('preparing_page', 1))
    ready_page     = Paginator(ready_orders,     _per_page).get_page(request.GET.get('ready_page',     1))
    served_page    = Paginator(served_orders,    _per_page).get_page(request.GET.get('served_page',    1))
    cancelled_page = Paginator(cancelled_orders, _per_page).get_page(request.GET.get('cancelled_page', 1))

    context = {
        'pending_orders':   pending_page,
        'confirmed_orders': confirmed_page,
        'preparing_orders': preparing_page,
        'ready_orders':     ready_page,
        'served_orders':    served_page,
        'cancelled_orders': cancelled_page,
        'pending_page_param':   'pending_page',
        'confirmed_page_param': 'confirmed_page',
        'preparing_page_param': 'preparing_page',
        'ready_page_param':     'ready_page',
        'served_page_param':    'served_page',
        'cancelled_page_param': 'cancelled_page',
        'pending_count':   pending_count,
        'confirmed_count': confirmed_count,
        'preparing_count': preparing_count,
        'ready_count':     ready_count,
        'served_count':    served_count,
        'cancelled_count': cancelled_count,
        'total_count':     total_count,
        'period':          period,
        'restaurant_name': request.user.get_restaurant_name() if not request.user.is_administrator() else "All Restaurants",
        **restaurant_context,
    }

    return render(request, 'admin_panel/manage_orders.html', context)


@login_required
def manage_tables(request):
    """Table management view"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator/Owner privileges required.")
        return redirect('restaurant:home')

    try:
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        view_all_restaurants = restaurant_context['view_all_restaurants']
        
        # Get filter parameters
        search_query = request.GET.get('search', '').strip()
        restaurant_filter = request.GET.get('restaurant', '').strip()
        status_filter = request.GET.get('status', '').strip()
        
        # Determine which restaurant to filter by
        target_restaurant = None
        if restaurant_filter:
            try:
                from restaurant.models_restaurant import Restaurant
                target_restaurant = Restaurant.objects.get(id=restaurant_filter)
                # Security: verify user has access to this restaurant
                if not request.user.is_administrator():
                    if not restaurant_context['accessible_restaurants'].filter(id=target_restaurant.id).exists():
                        target_restaurant = None
            except Restaurant.DoesNotExist:
                target_restaurant = None
        elif current_restaurant and not view_all_restaurants:
            target_restaurant = current_restaurant
        
        if target_restaurant:
            # Filter tables by specific selected restaurant
            if target_restaurant.is_main_restaurant:
                # Main restaurant: show tables assigned to it OR owned by main owner with no restaurant
                tables = TableInfo.objects.filter(
                    Q(restaurant=target_restaurant) |
                    Q(owner=target_restaurant.main_owner, restaurant__isnull=True)
                )
            else:
                # Branch restaurant: only show tables specifically assigned to this branch
                # OR created by the branch owner (not main owner)
                branch_query = Q(restaurant=target_restaurant)
                if target_restaurant.branch_owner:
                    branch_query |= Q(owner=target_restaurant.branch_owner, restaurant__isnull=True)
                tables = TableInfo.objects.filter(branch_query)
        elif request.user.is_administrator():
            # Administrator sees all tables
            tables = TableInfo.objects.all()
        else:
            # Get tables from all accessible restaurants
            accessible_restaurants = restaurant_context['accessible_restaurants']
            
            if accessible_restaurants.exists():
                table_query = Q()
                for restaurant in accessible_restaurants:
                    table_query |= (
                        Q(restaurant=restaurant) |
                        Q(owner=restaurant.main_owner, restaurant__isnull=True) |
                        Q(owner=restaurant.branch_owner, restaurant__isnull=True)
                    )
                tables = TableInfo.objects.filter(table_query)
            else:
                tables = TableInfo.objects.none()
        
        # Apply search filter
        if search_query:
            tables = tables.filter(
                Q(tbl_no__icontains=search_query) |
                Q(capacity__icontains=search_query)
            )
        
        # Apply status filter
        if status_filter:
            if status_filter == 'available':
                tables = tables.filter(is_available=True)
            elif status_filter == 'occupied':
                tables = tables.filter(is_available=False)
            
        # Custom sorting to handle T01, T02, T10, T011 properly
        tables = sorted(tables, key=lambda x: (len(x.tbl_no), x.tbl_no))
        
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    # Pagination
    try:
        per_page = int(request.GET.get('per_page', 5))
    except (ValueError, TypeError):
        per_page = 5
    paginator = Paginator(tables, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'tables': page_obj.object_list,
        'page_obj': page_obj,
        'restaurant_name': current_restaurant.name if current_restaurant else (request.user.get_restaurant_name() if not request.user.is_administrator() else "All Restaurants"),
        **restaurant_context,  # Include restaurant context
    }

    return render(request, 'admin_panel/manage_tables.html', context)


@login_required
def manage_categories(request):
    """Category management view"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator/Owner privileges required.")
        return redirect('restaurant:home')

    # Initialize variables
    main_categories = MainCategory.objects.none()
    subcategories = SubCategory.objects.none()

    try:
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        view_all_restaurants = restaurant_context['view_all_restaurants']
        
        if current_restaurant and not view_all_restaurants:
            # Filter categories by specific selected restaurant
            if current_restaurant.is_main_restaurant:
                # Main restaurant: show categories assigned to it OR owned by main owner with no restaurant
                main_categories = MainCategory.objects.filter(
                    Q(restaurant=current_restaurant) | 
                    Q(owner=current_restaurant.main_owner, restaurant__isnull=True)
                ).order_by('name')
            else:
                # Branch restaurant: only show categories specifically assigned to this branch
                # OR categories created by the branch owner (not main owner)
                branch_query = Q(restaurant=current_restaurant)
                if current_restaurant.branch_owner:
                    branch_query |= Q(owner=current_restaurant.branch_owner, restaurant__isnull=True)
                    
                main_categories = MainCategory.objects.filter(branch_query).order_by('name')
            
            subcategories = SubCategory.objects.filter(
                Q(main_category__in=main_categories)
            ).order_by('main_category__name', 'name')
        elif request.user.is_administrator():
            # Administrator sees all categories
            main_categories = MainCategory.objects.all().order_by('name')
            subcategories = SubCategory.objects.all().order_by('main_category__name', 'name')
        else:
            # Get user's accessible restaurants
            accessible_restaurants = restaurant_context['accessible_restaurants']
            
            if accessible_restaurants.exists():
                # Build query for accessible restaurants and legacy owner field
                restaurant_query = Q()
                owner_query = Q()
                
                for restaurant in accessible_restaurants:
                    restaurant_query |= Q(restaurant=restaurant)
                    if restaurant.main_owner:
                        owner_query |= Q(owner=restaurant.main_owner, restaurant__isnull=True)
                    if restaurant.branch_owner:
                        owner_query |= Q(owner=restaurant.branch_owner, restaurant__isnull=True)
                
                main_categories = MainCategory.objects.filter(
                    restaurant_query | owner_query
                ).order_by('name')
                
                subcategories = SubCategory.objects.filter(
                    Q(main_category__in=main_categories)
                ).order_by('main_category__name', 'name')
            else:
                main_categories = MainCategory.objects.none()
                subcategories = SubCategory.objects.none()
            
    except PermissionDenied:
        messages.error(request, 'You are not associated with any restaurant.')
        return redirect('restaurant:home')

    # Apply filters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()
    restaurant_filter = request.GET.get('restaurant', '').strip()

    if search_query:
        main_categories = main_categories.filter(name__icontains=search_query)

    if status_filter:
        if status_filter == 'active':
            main_categories = main_categories.filter(is_active=True)
        elif status_filter == 'inactive':
            main_categories = main_categories.filter(is_active=False)

    if restaurant_filter:
        from restaurant.models_restaurant import Restaurant
        try:
            filter_restaurant = Restaurant.objects.get(id=restaurant_filter)
            # Filter categories by restaurant OR by owner (for legacy categories without restaurant)
            if filter_restaurant.is_main_restaurant:
                # Main restaurant: include categories assigned to it OR owned by main_owner with no restaurant
                main_categories = main_categories.filter(
                    Q(restaurant=filter_restaurant) |
                    Q(owner=filter_restaurant.main_owner, restaurant__isnull=True)
                )
            else:
                # Branch: include categories assigned to this branch OR owned by branch_owner with no restaurant
                branch_query = Q(restaurant=filter_restaurant)
                if filter_restaurant.branch_owner:
                    branch_query |= Q(owner=filter_restaurant.branch_owner, restaurant__isnull=True)
                main_categories = main_categories.filter(branch_query)
        except Restaurant.DoesNotExist:
            pass

    # Annotate counts to avoid N+1 in template, and select_related for subcategory FK accesses
    main_categories = main_categories.annotate(subcategory_count=Count('subcategories', distinct=True))
    subcategories = subcategories.select_related(
        'main_category', 'main_category__restaurant', 'main_category__owner'
    ).annotate(product_count=Count('products', distinct=True))

    # Pagination for main categories
    try:
        per_page = int(request.GET.get('per_page', 20))
    except (ValueError, TypeError):
        per_page = 20
    paginator = Paginator(main_categories, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'main_categories': page_obj.object_list,
        'page_obj': page_obj,
        'subcategories': subcategories,
        'restaurant_name': current_restaurant.name if current_restaurant else (request.user.get_restaurant_name() if not request.user.is_administrator() else "All Restaurants"),
        **restaurant_context,  # Include restaurant context
    }

    return render(request, 'admin_panel/manage_categories.html', context)


@login_required
@require_POST
@transaction.atomic
def add_main_category(request):
    """Add a new main category"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        # Get restaurant context to determine where to save the category
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        image = request.FILES.get('image')  # Handle image upload
        restaurant_id = request.POST.get('restaurant_id', '').strip()

        if not name:
            return JsonResponse({'success': False, 'message': 'Category name is required'})
        
        # Get target restaurant from form if provided
        from restaurant.models_restaurant import Restaurant
        target_restaurant = None
        
        if restaurant_id:
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
                # Verify user has permission to assign to this restaurant
                accessible_restaurants = restaurant_context['accessible_restaurants']
                if not accessible_restaurants.filter(id=target_restaurant.id).exists():
                    return JsonResponse({'success': False, 'message': 'You do not have permission to add categories to this restaurant'})
            except Restaurant.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Selected restaurant not found'})
        else:
            # Use current restaurant context if no restaurant specified
            target_restaurant = current_restaurant
            
        if not target_restaurant:
            return JsonResponse({'success': False, 'message': 'No restaurant context available. Please select a restaurant.'})

        # Check for duplicate names within the target restaurant's categories
        if target_restaurant.is_main_restaurant:
            # For main restaurant, check categories for this restaurant or main owner
            existing_query = Q(name__iexact=name) & (
                Q(restaurant=target_restaurant) | 
                Q(owner=target_restaurant.main_owner, restaurant__isnull=True)
            )
        else:
            # For branch, check categories for this specific branch
            existing_query = Q(name__iexact=name) & (
                Q(restaurant=target_restaurant) | 
                Q(owner=target_restaurant.branch_owner, restaurant__isnull=True)
            )
            
        if MainCategory.objects.filter(existing_query).exists():
            return JsonResponse({'success': False, 'message': f'Category with this name already exists in {target_restaurant.name}'})

        if image:
            try:
                validate_image_file(image)
            except ValidationError as e:
                return JsonResponse({'success': False, 'message': str(e.message)})

        # Create category assigned to target restaurant
        category = MainCategory.objects.create(
            name=name,
            description=description,
            image=image,
            restaurant=target_restaurant,  # Use restaurant field for new system
            owner=target_restaurant.branch_owner or target_restaurant.main_owner  # Backward compatibility
        )
        
        success_message = f'Main category "{name}" added successfully to {target_restaurant.name}'

        return JsonResponse({
            'success': True,
            'message': 'Main category added successfully',
            'category': {
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'is_active': category.is_active,
                'image_url': category.image.url if category.image else None
            }
        })

    except Exception as e:
        logger.error(f'Error adding main category: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to add category. Please try again.'})


@login_required
@require_POST
@transaction.atomic
def edit_main_category(request, category_id):
    """Edit an existing main category"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        # Import restaurant context utilities
        from admin_panel.restaurant_utils import get_restaurant_context
        from restaurant.models import Restaurant
        
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get the category and verify access
        try:
            category = MainCategory.objects.get(id=category_id)
            
            # Verify user has access to this category's restaurant
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if not request.user.is_administrator():
                if category.restaurant:
                    if not accessible_restaurants.filter(id=category.restaurant.id).exists():
                        return JsonResponse({'success': False, 'message': 'You do not have permission to edit this category'})
                else:
                    # Legacy category (no restaurant FK): check via owner
                    owner_filter = get_owner_filter(request.user)
                    if category.owner_id != (owner_filter.id if owner_filter else None):
                        return JsonResponse({'success': False, 'message': 'You do not have permission to edit this category'})
                
        except MainCategory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Category not found'})
            
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        image = request.FILES.get('image')  # Handle image upload
        restaurant_id = request.POST.get('restaurant')  # Get restaurant_id from form

        if not name:
            return JsonResponse({'success': False, 'message': 'Category name is required'})

        # Determine if restaurant is being changed
        current_category_restaurant_id = category.restaurant.id if category.restaurant else None
        
        if restaurant_id and str(current_category_restaurant_id) != str(restaurant_id):
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
                
                # Validate user has permission for this restaurant
                accessible_restaurants = restaurant_context.get('accessible_restaurants')
                if accessible_restaurants is None or not accessible_restaurants.filter(id=target_restaurant.id).exists():
                    return JsonResponse({'success': False, 'message': 'You do not have permission to assign categories to this restaurant'})
                
                # Determine owner based on restaurant type
                if target_restaurant.is_main_restaurant:
                    new_owner = target_restaurant.main_owner
                else:
                    new_owner = target_restaurant.branch_owner or target_restaurant.main_owner
                
                # Check for duplicates in target restaurant
                _mcq_move = Q(restaurant=target_restaurant)
                if new_owner:
                    _mcq_move |= (Q(owner=new_owner) | Q(restaurant__main_owner=new_owner) | Q(restaurant__branch_owner=new_owner))
                if MainCategory.objects.filter(_mcq_move, name__iexact=name).exclude(id=category_id).exists():
                    return JsonResponse({'success': False, 'message': f'Category with this name already exists in {target_restaurant.name}'})
                    
                category.owner = new_owner
                category.restaurant = target_restaurant
                    
            except Restaurant.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Selected restaurant does not exist'})
        else:
            # Same restaurant or no change, check duplicates with current owner
            if category.owner:
                _mcq_same = (
                    Q(owner=category.owner) |
                    Q(restaurant__main_owner=category.owner) |
                    Q(restaurant__branch_owner=category.owner)
                )
            else:
                _mcq_same = Q(restaurant=category.restaurant)
            if MainCategory.objects.filter(_mcq_same, name__iexact=name).exclude(id=category_id).exists():
                return JsonResponse({'success': False, 'message': 'Category with this name already exists in this restaurant'})

        category.name = name
        category.description = description
        if image:  # Only update image if a new one is uploaded
            try:
                validate_image_file(image)
            except ValidationError as e:
                return JsonResponse({'success': False, 'message': str(e.message)})
            category.image = image
        category.save()

        return JsonResponse({
            'success': True,
            'message': 'Main category updated successfully',
            'category': {
                'id': category.id,
                'name': category.name,
                'description': category.description,
                'is_active': category.is_active,
                'image_url': category.image.url if category.image else None
            }
        })

    except Exception as e:
        logger.error(f'Error editing main category: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to update category. Please try again.'})


@login_required
@require_POST
@transaction.atomic
def delete_main_category(request, category_id):
    """Delete a main category"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        # Import restaurant context utilities
        from admin_panel.restaurant_utils import get_restaurant_context
        
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get the category and verify access
        try:
            category = MainCategory.objects.get(id=category_id)
            
            # Verify user has access to this category's restaurant
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if not request.user.is_administrator():
                if category.restaurant:
                    if not accessible_restaurants.filter(id=category.restaurant.id).exists():
                        return JsonResponse({'success': False, 'message': 'You do not have permission to delete this category'})
                else:
                    # Legacy category (no restaurant FK): check via owner
                    owner_filter = get_owner_filter(request.user)
                    if category.owner_id != (owner_filter.id if owner_filter else None):
                        return JsonResponse({'success': False, 'message': 'You do not have permission to delete this category'})
                
        except MainCategory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Category not found'})
        
        category_name = category.name
        subcategory_count = category.subcategories.count()

        # Gather all products in this category
        from restaurant.models import Product as MenuProduct
        from orders.models import OrderItem
        category_products = MenuProduct.objects.filter(main_category=category)

        # Snapshot product IDs now so the subquery stays valid through all steps
        product_ids = list(category_products.values_list('id', flat=True))

        # Nullify all PROTECT FK references before deletion
        _cleanup_product_dependencies(product_ids)

        category_products.delete()
        category.subcategories.all().delete()
        category.delete()

        if subcategory_count > 0:
            message = f'Main category "{category_name}" and its {subcategory_count} subcategories deleted successfully'
        else:
            message = f'Main category "{category_name}" deleted successfully'

        return JsonResponse({
            'success': True,
            'message': message
        })

    except Exception as e:
        logger.error(f'Error deleting main category: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to delete category. Please try again.'})


@login_required
@require_POST
def toggle_main_category(request, category_id):
    """Toggle main category active status"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        # Import restaurant context utilities
        from admin_panel.restaurant_utils import get_restaurant_context
        
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get the category and verify access
        try:
            category = MainCategory.objects.get(id=category_id)
            
            # Verify user has access to this category's restaurant
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if not request.user.is_administrator():
                if category.restaurant:
                    if not accessible_restaurants.filter(id=category.restaurant.id).exists():
                        return JsonResponse({'success': False, 'message': 'You do not have permission to toggle this category'})
                else:
                    # Legacy category (no restaurant FK): check via owner
                    owner_filter = get_owner_filter(request.user)
                    if category.owner_id != (owner_filter.id if owner_filter else None):
                        return JsonResponse({'success': False, 'message': 'You do not have permission to toggle this category'})
                
        except MainCategory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Category not found'})
        
        category.is_active = not category.is_active
        category.save()

        status = 'activated' if category.is_active else 'deactivated'

        return JsonResponse({
            'success': True,
            'message': f'Main category "{category.name}" {status} successfully',
            'is_active': category.is_active
        })

    except Exception as e:
        logger.error(f'Error toggling main category: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to toggle category. Please try again.'})


@login_required
@require_POST
@transaction.atomic
def add_subcategory(request):
    """Add a new subcategory"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        # Get restaurant context instead of owner filter
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        
        main_category_id = request.POST.get('main_category', '').strip()
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()

        # Validate required fields and ensure main_category_id is numeric
        if not main_category_id or not name:
            return JsonResponse({'success': False, 'message': 'Main category and subcategory name are required'})
        
        try:
            main_category_id = int(main_category_id)
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'message': 'Invalid main category selected'})

        # Get main category with restaurant context filtering
        if current_restaurant:
            if current_restaurant.is_main_restaurant:
                # Main restaurant: show categories assigned to it OR owned by main owner with no restaurant
                try:
                    main_category = MainCategory.objects.get(
                        Q(id=main_category_id) & (
                            Q(restaurant=current_restaurant) | 
                            Q(owner=current_restaurant.main_owner, restaurant__isnull=True)
                        )
                    )
                except MainCategory.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Main category not found or access denied'})
            else:
                # Branch restaurant: only show categories specifically assigned to this branch
                branch_query = Q(restaurant=current_restaurant)
                if current_restaurant.branch_owner:
                    branch_query |= Q(owner=current_restaurant.branch_owner, restaurant__isnull=True)
                
                try:
                    main_category = MainCategory.objects.get(
                        Q(id=main_category_id) & branch_query
                    )
                except MainCategory.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Main category not found or access denied'})
        elif request.user.is_administrator():
            try:
                main_category = MainCategory.objects.get(id=main_category_id)
            except MainCategory.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Main category not found'})
        else:
            return JsonResponse({'success': False, 'message': 'No restaurant context found'})

        if SubCategory.objects.filter(main_category=main_category, name__iexact=name).exists():
            return JsonResponse({'success': False, 'message': 'Subcategory with this name already exists in the selected main category'})

        subcategory = SubCategory.objects.create(
            main_category=main_category,
            name=name,
            description=description
        )

        return JsonResponse({
            'success': True,
            'message': 'Subcategory added successfully',
            'subcategory': {
                'id': subcategory.id,
                'name': subcategory.name,
                'description': subcategory.description,
                'main_category': subcategory.main_category.name,
                'is_active': subcategory.is_active
            }
        })

    except Exception as e:
        logger.error(f'Error adding subcategory: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to add subcategory. Please try again.'})


@login_required
@require_POST
@transaction.atomic
def edit_subcategory(request, subcategory_id):
    """Edit an existing subcategory"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get subcategory with owner filtering
        if owner_filter:
            _scq = (
                Q(main_category__owner=owner_filter) |
                Q(main_category__restaurant__main_owner=owner_filter) |
                Q(main_category__restaurant__branch_owner=owner_filter)
            )
            subcategory = get_object_or_404(SubCategory.objects.filter(_scq), id=subcategory_id)
        else:
            subcategory = get_object_or_404(SubCategory, id=subcategory_id)
            
        main_category_id = request.POST.get('main_category')
        name = request.POST.get('name')
        description = request.POST.get('description', '')

        if not main_category_id or not name:
            return JsonResponse({'success': False, 'message': 'Main category and subcategory name are required'})

        # Get main category with owner filtering
        if owner_filter:
            _mcq = (
                Q(owner=owner_filter) |
                Q(restaurant__main_owner=owner_filter) |
                Q(restaurant__branch_owner=owner_filter)
            )
            main_category = get_object_or_404(MainCategory.objects.filter(_mcq), id=main_category_id)
        else:
            main_category = get_object_or_404(MainCategory, id=main_category_id)

        if SubCategory.objects.filter(main_category=main_category, name__iexact=name).exclude(id=subcategory_id).exists():
            return JsonResponse({'success': False, 'message': 'Subcategory with this name already exists in the selected main category'})

        subcategory.main_category = main_category
        subcategory.name = name
        subcategory.description = description
        subcategory.save()

        return JsonResponse({
            'success': True,
            'message': 'Subcategory updated successfully',
            'subcategory': {
                'id': subcategory.id,
                'name': subcategory.name,
                'description': subcategory.description,
                'main_category': subcategory.main_category.name,
                'is_active': subcategory.is_active
            }
        })

    except Exception as e:
        logger.error(f'Error editing subcategory: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to update subcategory. Please try again.'})


@login_required
@require_POST
@transaction.atomic
def delete_subcategory(request, subcategory_id):
    """Delete a subcategory"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        # Import restaurant context utilities
        from admin_panel.restaurant_utils import get_restaurant_context
        
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get the subcategory and verify access
        try:
            subcategory = SubCategory.objects.get(id=subcategory_id)
            
            # Verify user has access to this subcategory's restaurant (via main category)
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if not request.user.is_administrator():
                if subcategory.main_category.restaurant:
                    if not accessible_restaurants.filter(id=subcategory.main_category.restaurant.id).exists():
                        return JsonResponse({'success': False, 'message': 'You do not have permission to delete this subcategory'})
                else:
                    # Legacy category (no restaurant FK): check via owner
                    owner_filter = get_owner_filter(request.user)
                    if subcategory.main_category.owner_id != (owner_filter.id if owner_filter else None):
                        return JsonResponse({'success': False, 'message': 'You do not have permission to delete this subcategory'})
                
        except SubCategory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Subcategory not found'})
        
        subcategory_name = subcategory.name
        product_count = subcategory.products.count()

        # Nullify FK refs for products in this subcategory so SET_NULL cascade works cleanly
        product_ids = list(subcategory.products.values_list('id', flat=True))
        if product_ids:
            _cleanup_product_dependencies(product_ids)

        # Deleting the subcategory sets sub_category=NULL on its products (SET_NULL)
        subcategory.delete()

        if product_count > 0:
            message = f'Subcategory "{subcategory_name}" deleted successfully ({product_count} products unlinked)'
        else:
            message = f'Subcategory "{subcategory_name}" deleted successfully'

        return JsonResponse({
            'success': True,
            'message': message
        })

    except Exception as e:
        logger.error(f'Error deleting subcategory: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to delete subcategory. Please try again.'})


@login_required
@require_POST
def toggle_subcategory(request, subcategory_id):
    """Toggle subcategory active status"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        # Import restaurant context utilities
        from admin_panel.restaurant_utils import get_restaurant_context
        
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get the subcategory and verify access
        try:
            subcategory = SubCategory.objects.get(id=subcategory_id)
            
            # Verify user has access to this subcategory's restaurant (via main category)
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if not request.user.is_administrator():
                if subcategory.main_category.restaurant:
                    if not accessible_restaurants.filter(id=subcategory.main_category.restaurant.id).exists():
                        return JsonResponse({'success': False, 'message': 'You do not have permission to toggle this subcategory'})
                else:
                    # Legacy category (no restaurant FK): check via owner
                    owner_filter = get_owner_filter(request.user)
                    if subcategory.main_category.owner_id != (owner_filter.id if owner_filter else None):
                        return JsonResponse({'success': False, 'message': 'You do not have permission to toggle this subcategory'})
                
        except SubCategory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Subcategory not found'})
        
        subcategory.is_active = not subcategory.is_active
        subcategory.save()

        status = 'activated' if subcategory.is_active else 'deactivated'

        return JsonResponse({
            'success': True,
            'message': f'Subcategory "{subcategory.name}" {status} successfully',
            'is_active': subcategory.is_active
        })

    except Exception as e:
        logger.error(f'Error toggling subcategory: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to toggle subcategory. Please try again.'})


# Product CRUD API Views
@login_required
def get_subcategories(request, main_category_id):
    """Get subcategories for a main category"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Get restaurant context for proper filtering
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        view_all_restaurants = restaurant_context['view_all_restaurants']
        
        if current_restaurant and not view_all_restaurants:
            # Filter by specific selected restaurant
            if current_restaurant.is_main_restaurant:
                # Main restaurant: show categories assigned to it OR owned by main owner with no restaurant
                main_category = get_object_or_404(MainCategory, 
                    Q(id=main_category_id) & (
                        Q(restaurant=current_restaurant) | 
                        Q(owner=current_restaurant.main_owner, restaurant__isnull=True)
                    )
                )
            else:
                # Branch restaurant: only show categories specifically assigned to this branch
                # OR created by the branch owner (not main owner)
                branch_query = Q(restaurant=current_restaurant)
                if current_restaurant.branch_owner:
                    branch_query |= Q(owner=current_restaurant.branch_owner, restaurant__isnull=True)
                
                main_category = get_object_or_404(MainCategory,
                    Q(id=main_category_id) & branch_query
                )
        elif request.user.is_administrator():
            main_category = get_object_or_404(MainCategory, id=main_category_id)
        else:
            # Get user's accessible restaurants
            accessible_restaurants = restaurant_context['accessible_restaurants']
            
            if accessible_restaurants.exists():
                restaurant_query = Q()
                owner_query = Q()
                
                for restaurant in accessible_restaurants:
                    restaurant_query |= Q(restaurant=restaurant)
                    if restaurant.main_owner:
                        owner_query |= Q(owner=restaurant.main_owner, restaurant__isnull=True)
                    if restaurant.branch_owner:
                        owner_query |= Q(owner=restaurant.branch_owner, restaurant__isnull=True)
                
                main_category = get_object_or_404(MainCategory,
                    Q(id=main_category_id) & (restaurant_query | owner_query)
                )
            else:
                return JsonResponse({'success': False, 'message': 'No accessible restaurants'})
            
        subcategories = main_category.subcategories.filter(is_active=True).values('id', 'name')
        
        return JsonResponse({
            'success': True,
            'subcategories': list(subcategories)
        })
    except Exception as e:
        logger.error(f'Error getting subcategories: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to get subcategories. Please try again.'})


@login_required
@require_POST
@transaction.atomic
def bulk_delete_main_categories(request):
    """Bulk delete main categories"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        data = json.loads(request.body)
        category_ids = data.get('category_ids', [])
        
        if not category_ids:
            return JsonResponse({'success': False, 'message': 'No categories selected'})
        
        owner_filter = get_owner_filter(request.user)
        
        # Get categories with owner filtering
        if owner_filter:
            _mcq_bd = (
                Q(owner=owner_filter) |
                Q(restaurant__main_owner=owner_filter) |
                Q(restaurant__branch_owner=owner_filter)
            )
            categories = MainCategory.objects.filter(_mcq_bd, id__in=category_ids).distinct()
        else:
            categories = MainCategory.objects.filter(id__in=category_ids)
        
        if not categories.exists():
            return JsonResponse({'success': False, 'message': 'No valid categories found'})
        
        # Count related items before deletion
        category_names = list(categories.values_list('name', flat=True))
        total_subcategories = SubCategory.objects.filter(main_category__in=categories).count()
        total_products = Product.objects.filter(main_category__in=categories).count()

        # Nullify all PROTECT FK refs to these categories' products before deletion
        product_ids = list(Product.objects.filter(main_category__in=categories).values_list('id', flat=True))
        if product_ids:
            _cleanup_product_dependencies(product_ids)

        # Now delete in correct order: products → subcategories → categories
        deleted_count = categories.count()
        Product.objects.filter(id__in=product_ids).delete()
        SubCategory.objects.filter(main_category__in=categories).delete()
        categories.delete()
        
        # Build success message
        if deleted_count == 1:
            message = f'Main category "{category_names[0]}" deleted successfully'
        else:
            message = f'{deleted_count} main categories deleted successfully'
        
        if total_subcategories > 0:
            message += f' (including {total_subcategories} subcategories'
            if total_products > 0:
                message += f' and {total_products} products'
            message += ')'
        elif total_products > 0:
            message += f' (including {total_products} products)'

        return JsonResponse({
            'success': True,
            'message': message
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f'Error bulk deleting main categories: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to delete categories. Please try again.'})


@login_required
@require_POST
@transaction.atomic
def bulk_delete_subcategories(request):
    """Bulk delete subcategories"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})

    try:
        data = json.loads(request.body)
        subcategory_ids = data.get('subcategory_ids', [])
        
        if not subcategory_ids:
            return JsonResponse({'success': False, 'message': 'No subcategories selected'})
        
        owner_filter = get_owner_filter(request.user)
        
        # Get subcategories with owner filtering
        if owner_filter:
            subcategories = SubCategory.objects.filter(
                id__in=subcategory_ids
            ).filter(
                Q(main_category__owner=owner_filter) |
                Q(main_category__restaurant__main_owner=owner_filter) |
                Q(main_category__restaurant__branch_owner=owner_filter)
            ).distinct()
        else:
            subcategories = SubCategory.objects.filter(id__in=subcategory_ids)
        
        if not subcategories.exists():
            return JsonResponse({'success': False, 'message': 'No valid subcategories found'})
        
        # Count related products before deletion — 1 aggregate query instead of N
        subcategory_names = list(subcategories.values_list('name', flat=True))
        total_products = Product.objects.filter(sub_category__in=subcategories).count()
        
        # Nullify product deps so SET_NULL cascade on sub_category works cleanly
        product_ids = list(Product.objects.filter(sub_category__in=subcategories).values_list('id', flat=True))
        if product_ids:
            _cleanup_product_dependencies(product_ids)

        deleted_count = subcategories.count()
        subcategories.delete()

        # Build success message
        if deleted_count == 1:
            message = f'Subcategory "{subcategory_names[0]}" deleted successfully'
        else:
            message = f'{deleted_count} subcategories deleted successfully'

        if total_products > 0:
            message += f' ({total_products} products unlinked)'

        return JsonResponse({
            'success': True,
            'message': message
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'})
    except Exception as e:
        logger.error(f'Error bulk deleting subcategories: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to delete subcategories. Please try again.'})


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def add_product(request):
    """Add new product"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Get restaurant context for proper filtering
        from restaurant.models_restaurant import Restaurant
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get restaurant from form or use current restaurant
        restaurant_id = request.POST.get('restaurant')
        if restaurant_id:
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
                # Verify user has access to this restaurant
                accessible_restaurants = restaurant_context.get('accessible_restaurants')
                if accessible_restaurants is None or not accessible_restaurants.filter(id=target_restaurant.id).exists():
                    return JsonResponse({'success': False, 'message': 'You do not have permission to add products to this restaurant'})
                current_restaurant = target_restaurant
            except Restaurant.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Selected restaurant does not exist'})
        else:
            current_restaurant = restaurant_context['current_restaurant']
        
        # Get form data
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        main_category_id = request.POST.get('main_category', '').strip()
        sub_category_id = request.POST.get('sub_category', '').strip()
        price = request.POST.get('price', '').strip()
        stock = request.POST.get('available_in_stock', '').strip()
        prep_time = request.POST.get('preparation_time', '15').strip()
        is_available = request.POST.get('is_available') == 'on'
        image = request.FILES.get('image')
        
        # Validate required fields
        if not all([name, description, main_category_id, sub_category_id, price, stock]):
            return JsonResponse({'success': False, 'message': 'All required fields must be filled'})
        
        # Validate numeric fields
        try:
            main_category_id = int(main_category_id)
            sub_category_id = int(sub_category_id)
            prep_time = int(prep_time)
            price = float(price)
            if price < 0:
                raise ValueError('Price cannot be negative')
            stock = int(stock)
            if stock < 0:
                raise ValueError('Stock cannot be negative')
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'message': 'Invalid numeric values for category, preparation time, price, or stock'})
        
        # Get category objects with proper restaurant filtering
        if current_restaurant:
            if current_restaurant.is_main_restaurant:
                # Main restaurant: show categories assigned to it OR owned by main owner with no restaurant
                main_category = get_object_or_404(MainCategory, 
                    Q(id=main_category_id) & (
                        Q(restaurant=current_restaurant) | 
                        Q(owner=current_restaurant.main_owner, restaurant__isnull=True)
                    )
                )
            else:
                # Branch restaurant: only show categories specifically assigned to this branch
                # OR created by the branch owner (not main owner)
                branch_query = Q(restaurant=current_restaurant)
                if current_restaurant.branch_owner:
                    branch_query |= Q(owner=current_restaurant.branch_owner, restaurant__isnull=True)
                
                main_category = get_object_or_404(MainCategory,
                    Q(id=main_category_id) & branch_query
                )
                
            sub_category = get_object_or_404(SubCategory, 
                Q(id=sub_category_id) & Q(main_category=main_category)
            )
        elif request.user.is_administrator():
            main_category = get_object_or_404(MainCategory, id=main_category_id)
            sub_category = get_object_or_404(SubCategory, id=sub_category_id)
        else:
            return JsonResponse({'success': False, 'message': 'No restaurant context found'})
        
        # Get station field
        station = request.POST.get('station', 'kitchen')
        
        # Create product
        product = Product.objects.create(
            name=name,
            description=description,
            main_category=main_category,
            sub_category=sub_category,
            price=price,
            available_in_stock=stock,
            preparation_time=prep_time,
            is_available=is_available,
            station=station
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Product "{product.name}" added successfully',
            'product_id': product.id
        })
        
    except Exception as e:
        logger.error(f'Error adding product: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to add product. Please try again.'})


@login_required
def view_product(request, product_id):
    """Get product details for viewing"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get product with proper filtering for both restaurant and owner-based categories
        if owner_filter:
            # Build comprehensive query for accessible categories
            category_query = (
                Q(main_category__owner=owner_filter) |  # Categories owned directly
                Q(main_category__restaurant__main_owner=owner_filter) |  # Main restaurant categories
                Q(main_category__restaurant__branch_owner=owner_filter)  # Branch categories
            )
            product = get_object_or_404(Product.objects.filter(category_query), id=product_id)
        else:
            product = get_object_or_404(Product, id=product_id)
        
        html = render_to_string('admin_panel/product_detail.html', {
            'product': product
        })
        
        return JsonResponse({
            'success': True,
            'html': html
        })
        
    except Exception as e:
        logger.error(f'Error viewing product: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to load product details. Please try again.'})


@login_required
def edit_product(request, product_id):
    """Get product details for editing"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        owner_filter = get_owner_filter(request.user)
        
        # Get product with proper filtering for both restaurant and owner-based categories
        if owner_filter:
            # Build comprehensive query for accessible categories
            category_query = (
                Q(main_category__owner=owner_filter) |  # Categories owned directly
                Q(main_category__restaurant__main_owner=owner_filter) |  # Main restaurant categories
                Q(main_category__restaurant__branch_owner=owner_filter)  # Branch categories
            )
            product = get_object_or_404(Product.objects.filter(category_query), id=product_id)
            
            # Get accessible main categories with same comprehensive filtering
            main_categories = MainCategory.objects.filter(
                Q(owner=owner_filter) |
                Q(restaurant__main_owner=owner_filter) |
                Q(restaurant__branch_owner=owner_filter),
                is_active=True
            ).order_by('name')
        else:
            product = get_object_or_404(Product, id=product_id)
            main_categories = MainCategory.objects.filter(is_active=True).order_by('name')
        subcategories = product.main_category.subcategories.filter(is_active=True).order_by('name')
        
        html = render_to_string('admin_panel/product_edit_form.html', {
            'product': product,
            'main_categories': main_categories,
            'subcategories': subcategories
        }, request=request)
        
        return JsonResponse({
            'success': True,
            'html': html
        })
        
    except Exception as e:
        logger.error(f'Error loading product edit form: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to load product form. Please try again.'})


@login_required
@require_http_methods(["POST"])
def update_product(request, product_id):
    """Update product"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Access denied'})
        else:
            messages.error(request, "Access denied. Administrator privileges required.")
            return redirect('admin_panel:manage_products')
    
    try:
        # Get owner filter for permission check
        owner_filter = get_owner_filter(request.user)
        
        # Get product with proper filtering for both restaurant and owner-based categories
        if owner_filter:
            # Build comprehensive query for accessible categories
            # For managers: owner_filter is their owner (could be main_owner or branch_owner)
            category_query = (
                Q(main_category__owner=owner_filter) |  # Categories owned directly
                Q(main_category__restaurant__main_owner=owner_filter) |  # Main restaurant categories
                Q(main_category__restaurant__branch_owner=owner_filter)  # Branch categories where owner is branch_owner
            )
            product = get_object_or_404(Product.objects.filter(category_query), id=product_id)
        else:
            product = get_object_or_404(Product, id=product_id)
        
        # Update fields
        product.name = request.POST.get('name', product.name)
        product.description = request.POST.get('description', product.description)
        
        if request.POST.get('main_category'):
            # Validate category belongs to owner
            if owner_filter:
                main_category = get_object_or_404(
                    MainCategory.objects.filter(
                        Q(owner=owner_filter) | 
                        Q(restaurant__main_owner=owner_filter) | 
                        Q(restaurant__branch_owner=owner_filter)
                    ),
                    id=request.POST.get('main_category')
                )
                product.main_category = main_category
            else:
                product.main_category = get_object_or_404(MainCategory, id=request.POST.get('main_category'))
        
        if request.POST.get('sub_category'):
            product.sub_category = get_object_or_404(SubCategory, id=request.POST.get('sub_category'), main_category=product.main_category)
        
        if request.POST.get('price'):
            try:
                product.price = float(request.POST.get('price'))
            except (ValueError, TypeError):
                pass  # Keep existing price if invalid
        
        if request.POST.get('available_in_stock'):
            try:
                product.available_in_stock = int(request.POST.get('available_in_stock'))
            except (ValueError, TypeError):
                pass  # Keep existing stock if invalid
        
        if request.POST.get('preparation_time'):
            try:
                product.preparation_time = int(request.POST.get('preparation_time'))
            except (ValueError, TypeError):
                pass  # Keep existing prep time if invalid
        
        if request.POST.get('station'):
            product.station = request.POST.get('station')
        
        product.is_available = request.POST.get('is_available') == 'on'
        
        product.save()
        
        # Check if this is an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Product "{product.name}" updated successfully'
            })
        else:
            # Regular form submission - redirect with success message
            messages.success(request, f'Product "{product.name}" updated successfully')
            return redirect('admin_panel:manage_products')
        
    except Exception as e:
        logger.error(f"Error updating product: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Failed to update product. Please try again.'})
        else:
            messages.error(request, 'Error updating product. Please try again.')
            return redirect('admin_panel:manage_products')


@login_required
@require_http_methods(["POST"])
def toggle_product_availability(request, product_id):
    """Toggle product availability"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Get owner filter for permission check
        owner_filter = get_owner_filter(request.user)
        
        # Get product with proper filtering for both restaurant and owner-based categories
        if owner_filter:
            # Build comprehensive query for accessible categories
            category_query = (
                Q(main_category__owner=owner_filter) |  # Categories owned directly
                Q(main_category__restaurant__main_owner=owner_filter) |  # Main restaurant categories
                Q(main_category__restaurant__branch_owner=owner_filter)  # Branch categories
            )
            product = get_object_or_404(Product.objects.filter(category_query), id=product_id)
        else:
            product = get_object_or_404(Product, id=product_id)
        
        data = json.loads(request.body)
        
        product.is_available = data.get('is_available', not product.is_available)
        product.save()
        
        status = 'available' if product.is_available else 'unavailable'
        
        return JsonResponse({
            'success': True,
            'message': f'Product "{product.name}" is now {status}',
            'is_available': product.is_available
        })
        
    except Exception as e:
        logger.error(f'Error toggling product availability: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to toggle product. Please try again.'})


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def delete_product(request, product_id):
    """Delete product"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Get owner filter for permission check
        owner_filter = get_owner_filter(request.user)
        
        # Get product with proper filtering for both restaurant and owner-based categories
        if owner_filter:
            # Build comprehensive query for accessible categories
            category_query = (
                Q(main_category__owner=owner_filter) |  # Categories owned directly
                Q(main_category__restaurant__main_owner=owner_filter) |  # Main restaurant categories
                Q(main_category__restaurant__branch_owner=owner_filter)  # Branch categories
            )
            product = get_object_or_404(Product.objects.filter(category_query), id=product_id)
        else:
            product = get_object_or_404(Product, id=product_id)
        
        product_name = product.name
        _cleanup_product_dependencies([product.id])
        product.delete()

        return JsonResponse({
            'success': True,
            'message': f'Product "{product_name}" deleted successfully'
        })

    except Exception as e:
        logger.error(f'Error deleting product: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Failed to delete product. Please try again.'})


# ============================================================================
# USER MANAGEMENT CRUD OPERATIONS
# ============================================================================

def _can_manage_target_user(requesting_user, target_user):
    """Check if requesting_user has permission to manage target_user based on restaurant affiliation."""
    if requesting_user.is_administrator():
        return True
    if target_user.is_administrator() or target_user.is_superuser:
        return False
    if requesting_user.is_main_owner():
        from restaurant.models_restaurant import Restaurant as _Restaurant
        return (
            target_user.owner_id == requesting_user.id or
            _Restaurant.objects.filter(main_owner=requesting_user, branch_owner=target_user).exists()
        )
    elif requesting_user.is_branch_owner():
        return target_user.owner_id == requesting_user.id
    elif requesting_user.is_manager():
        if not requesting_user.owner_id:
            return False
        # Managers cannot manage peer managers (lateral attack prevention)
        if target_user.is_manager():
            return False
        return target_user.owner_id == requesting_user.owner_id
    elif requesting_user.is_owner():
        return target_user.owner_id == requesting_user.id
    return False


@login_required
@require_POST
def add_user(request):
    """Add a new user"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Get form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        role_name = request.POST.get('role', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        phone_number = request.POST.get('phone_number', '').strip()
        address = request.POST.get('address', '').strip()
        restaurant_id = request.POST.get('restaurant_id', '').strip()
        
        # Validation
        if not all([first_name, last_name, username, email, password, role_name]):
            return JsonResponse({'success': False, 'message': 'All required fields must be filled'})
        
        # Check role restrictions based on user hierarchy
        if not request.user.is_administrator():
            # Branch owners can ONLY create staff roles (not other branch_owner or main_owner)
            if request.user.is_branch_owner():
                if role_name in ['main_owner', 'branch_owner', 'owner', 'administrator']:
                    return JsonResponse({
                        'success': False, 
                        'message': 'Branch owners can only create staff members (kitchen, bar, cashier, etc.), not owner roles.'
                    })
            
            # Main owners can create branch_owner but NOT another main_owner or legacy owner
            elif request.user.is_main_owner():
                if role_name in ['main_owner', 'owner']:
                    return JsonResponse({
                        'success': False, 
                        'message': 'Cannot create main owner or legacy owner roles. You can only create branch owners and staff.'
                    })
                if role_name == 'administrator':
                    return JsonResponse({'success': False, 'message': 'Only administrators can create administrator accounts'})
            
            # Legacy owners can create any role except administrator and main_owner
            elif request.user.is_owner():
                if role_name in ['administrator', 'main_owner']:
                    return JsonResponse({'success': False, 'message': 'You cannot create administrator or main owner accounts'})
            
            # Managers can ONLY create regular staff roles (not owners, managers, or customers)
            elif request.user.is_manager():
                if role_name in ['administrator', 'main_owner', 'branch_owner', 'owner', 'manager', 'customer']:
                    return JsonResponse({
                        'success': False,
                        'message': 'Managers can only create staff roles (kitchen, bar, cashier, customer_care, service, buffet).'
                    })
        
        # Check if username or email already exists
        if User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'message': 'Username already exists'})
        
        if User.objects.filter(email=email).exists():
            return JsonResponse({'success': False, 'message': 'Email already exists'})
        
        # Get role
        try:
            role = Role.objects.get(name=role_name)
        except Role.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid role selected'})
        
        # Get restaurant context for proper user assignment
        from restaurant.models_restaurant import Restaurant
        target_restaurant = None
        
        if restaurant_id:
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
                # Verify user has permission to assign to this restaurant
                accessible_restaurants = Restaurant.get_accessible_restaurants(request.user)
                if not accessible_restaurants.filter(id=target_restaurant.id).exists():
                    return JsonResponse({'success': False, 'message': 'You do not have permission to assign users to this restaurant'})
            except Restaurant.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Selected restaurant not found'})
        else:
            # Get current restaurant context if no restaurant specified
            session_restaurant_id = request.session.get('selected_restaurant_id')
            restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
            target_restaurant = restaurant_context['current_restaurant']
            
            if not target_restaurant:
                return JsonResponse({'success': False, 'message': 'No restaurant context available. Please select a restaurant.'})
        
        # Validate role based on subscription plan (after target_restaurant is determined)
        if role_name == 'main_owner':
            # Only administrators can create main_owner
            if not request.user.is_administrator():
                return JsonResponse({
                    'success': False, 
                    'message': 'Only system administrators can create main owner accounts.'
                })
        
        if role_name in ['branch_owner', 'owner']:
            # Check if target restaurant has SINGLE plan - no branch owners allowed
            if not target_restaurant:
                return JsonResponse({
                    'success': False,
                    'message': 'Cannot create owner roles without a valid restaurant context.'
                })
            if target_restaurant.subscription_plan == 'SINGLE':
                return JsonResponse({
                    'success': False, 
                    'message': 'Cannot create owner roles for SINGLE plan restaurants. Upgrade to PRO plan to create branches with branch owners.'
                })
        
        # Create user
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            role=role,
            is_active=is_active,
            phone_number=phone_number,
            address=address
        )
        
        # Assign proper owner based on restaurant and role
        if role_name in ['main_owner', 'branch_owner']:
            # Owner roles don't need an owner assigned
            if role_name == 'main_owner' and target_restaurant:
                # Update restaurant main_owner if creating main owner
                target_restaurant.main_owner = user
                target_restaurant.save()
            elif role_name == 'branch_owner' and target_restaurant and not target_restaurant.is_main_restaurant:
                # Update restaurant branch_owner if creating branch owner
                target_restaurant.branch_owner = user
                target_restaurant.save()
        else:
            # Staff roles need to be assigned to the appropriate owner
            if target_restaurant:
                if target_restaurant.is_main_restaurant:
                    # Assign to main owner for main restaurant
                    user.owner = target_restaurant.main_owner
                else:
                    # Assign to branch owner for branches
                    user.owner = target_restaurant.branch_owner or target_restaurant.main_owner
                user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'User "{user.get_full_name() or user.username}" created successfully',
            'user': {
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'email': user.email,
                'role': role.get_name_display(),
                'is_active': user.is_active,
                'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M')
            }
        })
        
    except Exception as e:
        logger.error(f'Error creating user: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error creating user. Please try again.'})


@login_required
@require_http_methods(["GET"])
def get_user_data(request, user_id):
    """Get user data for editing"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Check if requesting user has permission to access this user
        if not _can_manage_target_user(request.user, user):
            return JsonResponse({'success': False, 'message': 'Access denied - you do not have permission to access this user'})
        
        return JsonResponse({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'role': user.role.name if user.role else '',
                'is_active': user.is_active,
                'phone_number': user.phone_number,
                'address': user.address,
                'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M'),
                'last_login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never'
            }
        })
        
    except Exception as e:
        logger.error(f'Error getting user data: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error retrieving user data. Please try again.'})


@login_required
@require_POST
def update_user(request, user_id):
    """Update an existing user"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Check if requesting user has permission to manage this user
        if not _can_manage_target_user(request.user, user):
            return JsonResponse({'success': False, 'message': 'Access denied - you do not have permission to manage this user'})
        
        # Get form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        role_name = request.POST.get('role', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        phone_number = request.POST.get('phone_number', '').strip()
        address = request.POST.get('address', '').strip()
        restaurant_id = request.POST.get('restaurant_id', '').strip()
        
        # Validation
        if not all([first_name, last_name, username, email, role_name]):
            return JsonResponse({'success': False, 'message': 'All required fields must be filled'})
        
        # Check if username or email already exists (excluding current user)
        if User.objects.filter(username=username).exclude(id=user_id).exists():
            return JsonResponse({'success': False, 'message': 'Username already exists'})
        
        if User.objects.filter(email=email).exclude(id=user_id).exists():
            return JsonResponse({'success': False, 'message': 'Email already exists'})
        
        # Get role
        try:
            role = Role.objects.get(name=role_name)
        except Role.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Invalid role selected'})
        
        # Check role restrictions based on user hierarchy (same as add_user)
        if not request.user.is_administrator():
            # Branch owners can ONLY assign staff roles (not other branch_owner or main_owner)
            if request.user.is_branch_owner():
                if role_name in ['main_owner', 'branch_owner', 'owner', 'administrator']:
                    return JsonResponse({
                        'success': False, 
                        'message': 'Branch owners can only assign staff roles (kitchen, bar, cashier, etc.), not owner roles.'
                    })
            
            # Main owners can assign branch_owner but NOT another main_owner or legacy owner
            elif request.user.is_main_owner():
                if role_name in ['main_owner', 'owner']:
                    return JsonResponse({
                        'success': False, 
                        'message': 'Cannot assign main owner or legacy owner roles. You can only assign branch owners and staff.'
                    })
                if role_name == 'administrator':
                    return JsonResponse({'success': False, 'message': 'Only administrators can assign administrator role'})
            
            # Legacy owners can assign any role except administrator and main_owner
            elif request.user.is_owner():
                if role_name in ['administrator', 'main_owner']:
                    return JsonResponse({'success': False, 'message': 'You cannot assign administrator or main owner roles'})
            
            # Managers can ONLY assign regular staff roles (not owners, managers, or customers)
            elif request.user.is_manager():
                if role_name in ['administrator', 'main_owner', 'branch_owner', 'owner', 'manager', 'customer']:
                    return JsonResponse({
                        'success': False,
                        'message': 'Managers can only assign staff roles (kitchen, bar, cashier, customer_care, service, buffet).'
                    })
        
        # Get restaurant context for subscription plan validation
        from restaurant.models_restaurant import Restaurant
        target_restaurant = None
        
        # Try to get target restaurant from request or user's current context
        if restaurant_id:
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
            except Restaurant.DoesNotExist:
                pass
        
        # If no restaurant specified, get from current session
        if not target_restaurant:
            session_restaurant_id = request.session.get('selected_restaurant_id')
            restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
            target_restaurant = restaurant_context.get('current_restaurant')
        
        # Validate role based on subscription plan (same as add_user)
        if role_name == 'main_owner':
            # Only administrators can assign main_owner
            if not request.user.is_administrator():
                return JsonResponse({
                    'success': False, 
                    'message': 'Only system administrators can assign main owner role.'
                })
        
        if role_name in ['branch_owner', 'owner']:
            # Check if target restaurant has SINGLE plan - no branch owners allowed
            if not target_restaurant:
                return JsonResponse({
                    'success': False,
                    'message': 'Cannot assign owner roles without a valid restaurant context.'
                })
            if target_restaurant.subscription_plan == 'SINGLE':
                return JsonResponse({
                    'success': False, 
                    'message': 'Cannot assign owner roles for SINGLE plan restaurants. Upgrade to PRO plan to assign branch owners.'
                })
        
        # Handle restaurant assignment if provided
        from restaurant.models_restaurant import Restaurant
        if restaurant_id:
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
                # Verify user has permission to assign to this restaurant
                accessible_restaurants = Restaurant.get_accessible_restaurants(request.user)
                if not accessible_restaurants.filter(id=target_restaurant.id).exists():
                    return JsonResponse({'success': False, 'message': 'You do not have permission to assign users to this restaurant'})
                
                # Update owner relationship for staff roles
                if role_name not in ['main_owner', 'branch_owner']:
                    if target_restaurant.is_main_restaurant:
                        user.owner = target_restaurant.main_owner
                    else:
                        user.owner = target_restaurant.branch_owner or target_restaurant.main_owner
            except Restaurant.DoesNotExist:
                pass  # Keep existing assignment if restaurant not found
        
        # Update user
        user.first_name = first_name
        user.last_name = last_name
        user.username = username
        user.email = email
        user.role = role
        user.is_active = is_active
        user.phone_number = phone_number
        user.address = address
        
        # Update password if provided
        if password:
            user.set_password(password)
        
        user.save()

        # Audit log: track user updates (especially role changes)
        try:
            from accounts.security_utils import log_security_event, get_client_ip
            log_security_event(
                event_type='staff_updated',
                user=request.user,
                description=f'User "{user.get_full_name() or user.username}" updated by {request.user.username} - role: {role.name}, active: {user.is_active}',
                ip_address=get_client_ip(request),
                extra_data={'updated_user_id': user_id, 'new_role': role.name, 'password_changed': bool(password)}
            )
        except Exception:
            pass

        return JsonResponse({
            'success': True,
            'message': f'User "{user.get_full_name() or user.username}" updated successfully',
            'user': {
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name(),
                'email': user.email,
                'role': role.get_name_display(),
                'is_active': user.is_active
            }
        })
        
    except Exception as e:
        logger.error(f'Error updating user: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error updating user. Please try again.'})


@login_required
@require_POST
def toggle_user_status(request, user_id):
    """Toggle user active/inactive status"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Check if requesting user has permission to manage this user
        if not _can_manage_target_user(request.user, user):
            return JsonResponse({'success': False, 'message': 'Access denied - you do not have permission to manage this user'})
        
        # Prevent deactivating self
        if user.id == request.user.id:
            return JsonResponse({'success': False, 'message': 'You cannot deactivate your own account'})
        
        # Prevent deactivating other administrators unless you're a superuser
        if user.is_administrator() and not request.user.is_superuser:
            return JsonResponse({'success': False, 'message': 'You cannot modify administrator accounts'})
        
        user.is_active = not user.is_active
        user.save()

        status = 'activated' if user.is_active else 'deactivated'

        # Audit log: account status changes are security-relevant
        try:
            from accounts.security_utils import log_security_event, get_client_ip
            log_security_event(
                event_type='staff_updated',
                user=request.user,
                description=f'User "{user.get_full_name() or user.username}" {status} by {request.user.username}',
                ip_address=get_client_ip(request),
                extra_data={'target_user_id': user_id, 'action': status}
            )
        except Exception:
            pass

        return JsonResponse({
            'success': True,
            'message': f'User "{user.get_full_name() or user.username}" {status} successfully',
            'is_active': user.is_active
        })
        
    except Exception as e:
        logger.error(f'Error toggling user status: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error updating user status. Please try again.'})


@login_required
@require_POST
@transaction.atomic
def delete_user(request, user_id):
    """Delete a user"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Check if requesting user has permission to delete this user
        if not _can_manage_target_user(request.user, user):
            return JsonResponse({'success': False, 'message': 'Access denied - you do not have permission to delete this user'})
        
        # Prevent deleting self
        if user.id == request.user.id:
            return JsonResponse({'success': False, 'message': 'You cannot delete your own account'})
        
        # Prevent deleting other administrators unless you're a superuser
        if user.is_administrator() and not request.user.is_superuser:
            return JsonResponse({'success': False, 'message': 'You cannot delete administrator accounts'})
        
        user_name = user.get_full_name() or user.username
        user_role = user.role.name if user.role else 'no-role'

        try:
            user.delete()
        except ProtectedError:
            return JsonResponse({
                'success': False,
                'message': (
                    f'Cannot delete user "{user_name}" because they have associated orders, '
                    'payments, or waste records. Deactivate the account instead.'
                )
            })

        # Audit log: user deletion is a critical security event
        try:
            from accounts.security_utils import log_security_event, get_client_ip
            log_security_event(
                event_type='staff_deactivated',
                user=request.user,
                description=f'User "{user_name}" (role: {user_role}) permanently deleted by {request.user.username}',
                ip_address=get_client_ip(request),
                extra_data={'deleted_user_id': user_id, 'deleted_username': user_name, 'deleted_role': user_role}
            )
        except Exception:
            pass

        return JsonResponse({
            'success': True,
            'message': f'User "{user_name}" deleted successfully'
        })

    except Exception as e:
        logger.error(f'Error deleting user: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error deleting user. Please try again.'})


# ============================================================================
# ROLE MANAGEMENT CRUD OPERATIONS
# ============================================================================

@login_required
@require_POST
def add_role(request):
    """Add a new role"""
    if not request.user.is_administrator():
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        
        # Validation
        if not name:
            return JsonResponse({'success': False, 'message': 'Role name is required'})
        
        # Check if role already exists
        if Role.objects.filter(name=name).exists():
            return JsonResponse({'success': False, 'message': 'Role already exists'})
        
        # Validate role name is in choices
        valid_roles = [choice[0] for choice in Role.ROLE_CHOICES]
        if name not in valid_roles:
            return JsonResponse({'success': False, 'message': 'Invalid role name'})
        
        # Create role
        role = Role.objects.create(
            name=name,
            description=description
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Role "{role.get_name_display()}" created successfully',
            'role': {
                'id': role.id,
                'name': role.name,
                'display_name': role.get_name_display(),
                'description': role.description,
                'created_at': role.created_at.strftime('%Y-%m-%d %H:%M')
            }
        })
        
    except Exception as e:
        logger.error(f'Error creating role: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error creating role. Please try again.'})


@login_required
@require_http_methods(["GET"])
def get_role_data(request, role_id):
    """Get role data for editing"""
    if not request.user.is_administrator():
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        role = get_object_or_404(Role, id=role_id)
        
        return JsonResponse({
            'success': True,
            'role': {
                'id': role.id,
                'name': role.name,
                'display_name': role.get_name_display(),
                'description': role.description,
                'created_at': role.created_at.strftime('%Y-%m-%d %H:%M'),
                'user_count': role.user_set.count()
            }
        })
        
    except Exception as e:
        logger.error(f'Error creating role: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error creating role. Please try again.'})


@login_required
@require_POST
def update_role(request, role_id):
    """Update an existing role"""
    if not request.user.is_administrator():
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        role = get_object_or_404(Role, id=role_id)
        
        description = request.POST.get('description', '').strip()
        
        # Update role (name cannot be changed for system roles)
        role.description = description
        role.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Role "{role.get_name_display()}" updated successfully',
            'role': {
                'id': role.id,
                'name': role.name,
                'display_name': role.get_name_display(),
                'description': role.description
            }
        })
        
    except Exception as e:
        logger.error(f'Error updating role: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error updating role. Please try again.'})


@login_required
@require_POST
def delete_role(request, role_id):
    """Delete a role"""
    if not request.user.is_administrator():
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        role = get_object_or_404(Role, id=role_id)
        
        # Check if role has users assigned
        if role.user_set.exists():
            return JsonResponse({
                'success': False, 
                'message': f'Cannot delete role "{role.get_name_display()}" because it has users assigned to it'
            })
        
        # Prevent deleting system roles
        system_roles = ['administrator', 'owner', 'customer_care', 'kitchen', 'customer']
        if role.name in system_roles:
            return JsonResponse({
                'success': False, 
                'message': f'Cannot delete system role "{role.get_name_display()}"'
            })
        
        role_name = role.get_name_display()
        role.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Role "{role_name}" deleted successfully'
        })
        
    except Exception as e:
        logger.error(f'Error deleting role: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error deleting role. Please try again.'})


@login_required
def edit_user(request, user_id):
    """Edit user view (GET request)"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator/Owner privileges required.")
        return redirect('admin_panel:manage_users')
    
    user = get_object_or_404(User, id=user_id)
    
    # Check if requesting user has permission to edit this user
    if not _can_manage_target_user(request.user, user):
        messages.error(request, "Access denied - you do not have permission to manage this user.")
        return redirect('admin_panel:manage_users')
    
    # Filter roles based on user permissions and subscription plan
    if request.user.is_administrator():
        roles = Role.objects.all()
    else:
        roles = Role.objects.exclude(name='administrator')
        
        # Get restaurant context to check subscription plan
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context.get('current_restaurant')
        
        # Apply role filtering based on user type and subscription plan
        if request.user.is_branch_owner():
            # Branch owners can ONLY assign staff roles
            roles = roles.exclude(name__in=['main_owner', 'branch_owner', 'owner'])
        elif request.user.is_main_owner():
            # Main owners cannot assign another main_owner, hide legacy 'owner' role
            roles = roles.exclude(name__in=['main_owner', 'owner'])
            # On SINGLE plan, also hide branch_owner
            if current_restaurant and current_restaurant.subscription_plan == 'SINGLE':
                roles = roles.exclude(name='branch_owner')
        elif request.user.is_manager():
            # Managers can ONLY assign regular staff roles (not owners, managers, or customers)
            roles = roles.exclude(name__in=['main_owner', 'branch_owner', 'owner', 'manager', 'customer'])
        elif current_restaurant and current_restaurant.subscription_plan == 'SINGLE':
            # SINGLE plan: Only allow staff roles (no owner roles)
            roles = roles.exclude(name__in=['main_owner', 'branch_owner', 'owner'])
    
    context = {
        'edit_user': user,
        'roles': roles,
    }
    
    return render(request, 'admin_panel/edit_user.html', context)


@login_required
def edit_role(request, role_id):
    """Edit role view (GET request)"""
    if not request.user.is_administrator():
        messages.error(request, "Access denied. Administrator privileges required.")
        return redirect('admin_panel:manage_users')
    
    role = get_object_or_404(Role, id=role_id)
    
    context = {
        'edit_role': role,
    }
    
    return render(request, 'admin_panel/edit_role.html', context)


# Table Management CRUD Views
@login_required
@require_http_methods(["POST"])
def add_table(request):
    """Add new table"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Get restaurant context to determine where to save the table
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        
        table_number = request.POST.get('tbl_no', '').strip()
        capacity = request.POST.get('capacity')
        is_available = request.POST.get('is_available') == 'on'
        restaurant_id = request.POST.get('restaurant_id', '').strip()
        
        if not table_number:
            return JsonResponse({'success': False, 'message': 'Table number is required'})
        
        # Get target restaurant from form if provided
        from restaurant.models_restaurant import Restaurant
        target_restaurant = None
        
        if restaurant_id:
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
                # Verify user has permission to assign to this restaurant
                accessible_restaurants = restaurant_context['accessible_restaurants']
                if not accessible_restaurants.filter(id=target_restaurant.id).exists():
                    return JsonResponse({'success': False, 'message': 'You do not have permission to add tables to this restaurant'})
            except Restaurant.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Selected restaurant not found'})
        else:
            # Use current restaurant context if no restaurant specified
            target_restaurant = current_restaurant
            
        if not target_restaurant:
            return JsonResponse({'success': False, 'message': 'No restaurant context available. Please select a restaurant.'})
        
        # Basic validation for table number (alphanumeric, max 10 chars)
        if len(table_number) > 10:
            return JsonResponse({'success': False, 'message': 'Table number must be 10 characters or less'})
        
        if not table_number.replace(' ', '').replace('-', '').isalnum():
            return JsonResponse({'success': False, 'message': 'Table number can only contain letters, numbers, spaces, and hyphens'})
        
        if not capacity or int(capacity) < 1:
            return JsonResponse({'success': False, 'message': 'Valid capacity is required'})
        
        # Check if table number already exists in target restaurant context
        if TableInfo.objects.filter(tbl_no=table_number, restaurant=target_restaurant).exists():
            return JsonResponse({'success': False, 'message': f'Table number already exists in {target_restaurant.name}'})
        
        # Create new table assigned to target restaurant
        table = TableInfo.objects.create(
            tbl_no=table_number,
            capacity=int(capacity),
            is_available=is_available,
            restaurant=target_restaurant,  # Use restaurant field for new system
            owner=target_restaurant.branch_owner or target_restaurant.main_owner  # Backward compatibility
        )
        
        success_message = f'Table {table_number} added successfully to {target_restaurant.name}'
        
        return JsonResponse({
            'success': True,
            'message': success_message
        })
        
    except Exception as e:
        logger.error(f'Error adding table: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error adding table. Please try again.'})


@login_required
def get_table(request):
    """Get table data for editing"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    table_id = request.GET.get('table_id')
    try:
        table = get_object_or_404(TableInfo, id=table_id)
        # Verify ownership for non-administrators
        if not request.user.is_administrator():
            from admin_panel.restaurant_utils import get_restaurant_context
            session_restaurant_id = request.session.get('selected_restaurant_id')
            restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if table.restaurant:
                if not accessible_restaurants.filter(id=table.restaurant.id).exists():
                    return JsonResponse({'success': False, 'message': 'Access denied'})
            else:
                # Legacy table (no restaurant FK): check via owner
                owner_filter = get_owner_filter(request.user)
                if table.owner_id != (owner_filter.id if owner_filter else None):
                    return JsonResponse({'success': False, 'message': 'Access denied'})
        return JsonResponse({
            'success': True,
            'table': {
                'id': table.id,
                'tbl_no': table.tbl_no,
                'capacity': table.capacity,
                'is_available': table.is_available,
                'restaurant_id': table.restaurant.id if table.restaurant else None
            }
        })
    except Exception as e:
        logger.error(f'Error getting table: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error retrieving table data. Please try again.'})


@login_required
@require_http_methods(["POST"])
def update_table(request):
    """Update table"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Import restaurant context utilities
        from admin_panel.restaurant_utils import get_restaurant_context
        from restaurant.models import Restaurant
        
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        owner_filter = get_owner_filter(request.user)
        
        table_id = request.POST.get('table_id')
        table_number = request.POST.get('tbl_no', '').strip()
        capacity = request.POST.get('capacity')
        is_available = request.POST.get('is_available') == 'on'
        restaurant_id = request.POST.get('restaurant')  # Get restaurant_id from form
        
        if not table_number:
            return JsonResponse({'success': False, 'message': 'Table number is required'})
        
        # Basic validation for table number (alphanumeric, max 10 chars)
        if len(table_number) > 10:
            return JsonResponse({'success': False, 'message': 'Table number must be 10 characters or less'})
        
        if not table_number.replace(' ', '').replace('-', '').isalnum():
            return JsonResponse({'success': False, 'message': 'Table number can only contain letters, numbers, spaces, and hyphens'})
        
        if not capacity or int(capacity) < 1:
            return JsonResponse({'success': False, 'message': 'Valid capacity is required'})
        
        # Get table and verify access
        try:
            table = TableInfo.objects.get(id=table_id)
            
            # Verify user has access to this table's restaurant
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if not request.user.is_administrator():
                if table.restaurant:
                    if not accessible_restaurants.filter(id=table.restaurant.id).exists():
                        return JsonResponse({'success': False, 'message': 'You do not have permission to edit this table'})
                else:
                    # Legacy table (no restaurant FK): check via owner
                    owner_filter_t = get_owner_filter(request.user)
                    if table.owner_id != (owner_filter_t.id if owner_filter_t else None):
                        return JsonResponse({'success': False, 'message': 'You do not have permission to edit this table'})
                
        except TableInfo.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Table not found'})
        
        # Determine if restaurant is being changed
        current_table_restaurant_id = table.restaurant.id if table.restaurant else None
        
        if restaurant_id and str(current_table_restaurant_id) != str(restaurant_id):
            # Restaurant is being changed
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
                
                # Validate user has permission for this restaurant
                accessible_restaurants = restaurant_context.get('accessible_restaurants')
                if accessible_restaurants is None or not accessible_restaurants.filter(id=target_restaurant.id).exists():
                    return JsonResponse({'success': False, 'message': 'You do not have permission to assign tables to this restaurant'})
                
                # Determine owner based on restaurant type
                if target_restaurant.is_main_restaurant:
                    new_owner = target_restaurant.main_owner
                else:
                    new_owner = target_restaurant.branch_owner or target_restaurant.main_owner
                
                # Check for duplicates in target restaurant
                if TableInfo.objects.filter(
                    Q(restaurant=target_restaurant) | Q(owner=new_owner),
                    tbl_no=table_number
                ).exclude(id=table_id).exists():
                    return JsonResponse({'success': False, 'message': f'Table {table_number} already exists in {target_restaurant.name}'})
                    
                table.owner = new_owner
                table.restaurant = target_restaurant
                    
            except Restaurant.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Selected restaurant does not exist'})
        else:
            # Same restaurant or no change, just check duplicates with current owner
            if owner_filter:
                _tiq_dup = (
                    Q(owner=owner_filter) |
                    Q(restaurant__main_owner=owner_filter) |
                    Q(restaurant__branch_owner=owner_filter)
                )
                if TableInfo.objects.filter(_tiq_dup, tbl_no=table_number).exclude(id=table_id).exists():
                    return JsonResponse({'success': False, 'message': 'Table number already exists in your restaurant'})
            else:
                if TableInfo.objects.filter(tbl_no=table_number).exclude(id=table_id).exists():
                    return JsonResponse({'success': False, 'message': 'Table number already exists'})
        
        # Update table
        table.tbl_no = table_number
        table.capacity = int(capacity)
        table.is_available = is_available
        table.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Table {table_number} updated successfully'
        })
        
    except Exception as e:
        logger.error(f'Error updating table: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error updating table. Please try again.'})


@login_required
@require_http_methods(["POST"])
def toggle_table_status(request):
    """Toggle table availability status"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Import restaurant context utilities
        from admin_panel.restaurant_utils import get_restaurant_context
        
        table_id = request.POST.get('table_id')
        action = request.POST.get('action')  # 'occupy' or 'free'
        
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get the table and verify access
        try:
            table = TableInfo.objects.get(id=table_id)
            
            # Verify user has access to this table's restaurant
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if not request.user.is_administrator():
                if table.restaurant:
                    if not accessible_restaurants.filter(id=table.restaurant.id).exists():
                        return JsonResponse({'success': False, 'message': 'You do not have permission to modify this table'})
                else:
                    # Legacy table (no restaurant FK): check via owner
                    owner_filter_t = get_owner_filter(request.user)
                    if table.owner_id != (owner_filter_t.id if owner_filter_t else None):
                        return JsonResponse({'success': False, 'message': 'You do not have permission to modify this table'})
                
        except TableInfo.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Table not found'})
        
        if action == 'occupy':
            table.is_available = False
            message = f'Table {table.tbl_no} marked as occupied'
        elif action == 'free':
            table.is_available = True
            message = f'Table {table.tbl_no} marked as available'
        else:
            return JsonResponse({'success': False, 'message': 'Invalid action'})
        
        table.save()
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        logger.error(f'Error toggling table status: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error updating table status. Please try again.'})


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def delete_table(request):
    """Delete table"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Import restaurant context utilities
        from admin_panel.restaurant_utils import get_restaurant_context
        
        table_id = request.POST.get('table_id')
        
        # Get restaurant context
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get the table and verify access
        try:
            table = TableInfo.objects.get(id=table_id)
            
            # Verify user has access to this table's restaurant
            accessible_restaurants = restaurant_context['accessible_restaurants']
            if not request.user.is_administrator():
                if table.restaurant:
                    if not accessible_restaurants.filter(id=table.restaurant.id).exists():
                        return JsonResponse({'success': False, 'message': 'You do not have permission to delete this table'})
                else:
                    # Legacy table (no restaurant FK): check via owner
                    owner_filter_t = get_owner_filter(request.user)
                    if table.owner_id != (owner_filter_t.id if owner_filter_t else None):
                        return JsonResponse({'success': False, 'message': 'You do not have permission to delete this table'})
                
        except TableInfo.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Table not found'})
        
        table_number = table.tbl_no
        # Detach orders that reference this table — table_info is now nullable (SET_NULL)
        Order.objects.filter(table_info=table).update(table_info=None)
        table.delete()

        return JsonResponse({
            'success': True,
            'message': f'Table {table_number} deleted successfully'
        })

    except Exception as e:
        logger.error(f'Error deleting table: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error deleting table. Please try again.'})


# Order Management CRUD Views
@login_required
def view_order(request, order_id):
    """View order details"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        owner_filter = get_owner_filter(request.user)
        order_qs = Order.objects.select_related('ordered_by', 'table_info').prefetch_related('order_items__product')
        if owner_filter:
            _oq_vod = (
                Q(table_info__owner=owner_filter) |
                Q(table_info__restaurant__main_owner=owner_filter) |
                Q(table_info__restaurant__branch_owner=owner_filter)
            )
            order = get_object_or_404(order_qs.filter(_oq_vod), id=order_id)
        else:
            order = get_object_or_404(order_qs, id=order_id)
        order_items = order.order_items.all()
        
        context = {
            'order': order,
            'order_items': order_items,
        }
        
        html = render_to_string('admin_panel/order_details.html', context, request=request)
        return JsonResponse({'success': True, 'html': html})
        
    except Exception as e:
        logger.error(f'Error viewing order: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error viewing order. Please try again.'})


@login_required
@require_http_methods(["POST"])
def update_order_status(request):
    """Update order status"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        order_id = request.POST.get('order_id')
        new_status = request.POST.get('status')
        
        if not order_id or not new_status:
            return JsonResponse({'success': False, 'message': 'Order ID and status are required'})
        
        owner_filter = get_owner_filter(request.user)
        if owner_filter:
            _oq_uos = (
                Q(table_info__owner=owner_filter) |
                Q(table_info__restaurant__main_owner=owner_filter) |
                Q(table_info__restaurant__branch_owner=owner_filter)
            )
            order = get_object_or_404(Order.objects.filter(_oq_uos), id=order_id)
        else:
            order = get_object_or_404(Order, id=order_id)
        
        # Validate status
        valid_statuses = [choice[0] for choice in Order.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'message': 'Invalid status'})
        
        old_status = order.status
        order.status = new_status
        
        # If confirming order, set confirmed_by
        if new_status == 'confirmed' and not order.confirmed_by:
            order.confirmed_by = request.user
        
        order.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Order #{order.order_number} status updated from {old_status} to {new_status}'
        })
        
    except Exception as e:
        logger.error(f'Error updating order status: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error updating order status. Please try again.'})


@login_required
def add_order(request):
    """Add new order"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator privileges required.")
        return redirect('restaurant:home')
    
    if request.method == 'POST':
        try:
            owner_filter = get_owner_filter(request.user)
            
            table_id = request.POST.get('table_id')
            customer_id = request.POST.get('customer_id')
            special_instructions = request.POST.get('special_instructions', '')
            
            if not table_id or not customer_id:
                return JsonResponse({'success': False, 'message': 'Table and customer are required'})
            
            # Get table and customer with owner filtering
            if owner_filter:
                _tq_ao = (
                    Q(owner=owner_filter) |
                    Q(restaurant__main_owner=owner_filter) |
                    Q(restaurant__branch_owner=owner_filter)
                )
                table = get_object_or_404(TableInfo.objects.filter(_tq_ao).distinct(), id=table_id)
                customer = get_object_or_404(User, id=customer_id, owner=owner_filter)
            else:
                table = get_object_or_404(TableInfo, id=table_id)
                customer = get_object_or_404(User, id=customer_id)
            
            # Generate order number
            import random
            import string
            order_number = ''.join(random.choices(string.digits, k=8))
            while Order.objects.filter(order_number=order_number).exists():
                order_number = ''.join(random.choices(string.digits, k=8))
            
            # Create order
            order = Order.objects.create(
                order_number=order_number,
                table_info=table,
                ordered_by=customer,
                special_instructions=special_instructions,
                status='pending'
            )
            
            # ✨ SERVER-SIDE AUTO-PRINT
            try:
                from orders.printing import auto_print_order
                print_result = auto_print_order(order)
                print_message = f'Order #{order_number} created successfully'
                if print_result['kot_printed']:
                    print_message += ' | KOT printed!'
                if print_result['bot_printed']:
                    print_message += ' | BOT printed!'
            except Exception as print_error:
                logger.warning(f"Auto-print error: {print_error}")
                print_message = f'Order #{order_number} created (auto-print unavailable)'
            
            return JsonResponse({
                'success': True,
                'message': print_message
            })
            
        except Exception as e:
            logger.error(f'Error adding order: {str(e)}')
            return JsonResponse({'success': False, 'message': 'Error adding order. Please try again.'})
    
    # GET request - redirect to manage orders page (order creation uses the modal on that page)
    return redirect('admin_panel:manage_orders')


@login_required
def edit_order(request, order_id):
    """Edit order"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator privileges required.")
        return redirect('restaurant:home')
    
    # Get owner filter for multi-tenant support
    owner_filter = get_owner_filter(request.user)
    
    # Filter order by owner if owner, or get any order if administrator
    _order_qs = Order.objects.prefetch_related('order_items__product')
    if request.user.is_administrator():
        order = get_object_or_404(_order_qs, id=order_id)
    else:
        _oq_eo = (
            Q(table_info__owner=owner_filter) |
            Q(table_info__restaurant__main_owner=owner_filter) |
            Q(table_info__restaurant__branch_owner=owner_filter)
        )
        order = get_object_or_404(_order_qs.filter(_oq_eo), id=order_id)
    
    if request.method == 'POST':
        try:
            table_id = request.POST.get('table_id')
            customer_id = request.POST.get('customer_id')
            special_instructions = request.POST.get('special_instructions', '')
            status = request.POST.get('status')
            
            if table_id:
                # Filter table by owner for all non-admin users
                if owner_filter:
                    _tq_eo = (
                        Q(owner=owner_filter) |
                        Q(restaurant__main_owner=owner_filter) |
                        Q(restaurant__branch_owner=owner_filter)
                    )
                    table = get_object_or_404(TableInfo.objects.filter(_tq_eo).distinct(), id=table_id)
                else:
                    table = get_object_or_404(TableInfo, id=table_id)
                order.table_info = table
            
            if customer_id:
                if owner_filter:
                    order.ordered_by = get_object_or_404(User, id=customer_id, owner=owner_filter)
                else:
                    order.ordered_by = get_object_or_404(User, id=customer_id)
            
            order.special_instructions = special_instructions
            
            if status and status in [choice[0] for choice in Order.STATUS_CHOICES]:
                order.status = status
            
            order.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Order #{order.order_number} updated successfully'
            })
            
        except Exception as e:
            logger.error(f'Error editing order: {str(e)}')
            return JsonResponse({'success': False, 'message': 'Error editing order. Please try again.'})
    
    # GET request - show form
    # Filter tables by owner for all non-admin users
    if owner_filter:
        _tq_eo = (
            Q(owner=owner_filter) |
            Q(restaurant__main_owner=owner_filter) |
            Q(restaurant__branch_owner=owner_filter)
        )
        tables = TableInfo.objects.filter(_tq_eo).distinct()
        customers = User.objects.filter(role__name='Customer', owner=owner_filter)
    else:
        tables = TableInfo.objects.all()
        customers = User.objects.filter(role__name='Customer')
    
    context = {
        'order': order,
        'tables': tables,
        'customers': customers,
        'status_choices': Order.STATUS_CHOICES,
    }
    
    # Check if this is an AJAX request for modal content
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'admin_panel/edit_order_modal.html', context)
    else:
        return render(request, 'admin_panel/edit_order.html', context)


@login_required
@require_http_methods(["POST"])
@transaction.atomic
def delete_order(request, order_id):
    """Delete order"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    try:
        # Enforce restaurant tenancy — administrators may delete any order,
        # owners/managers may only delete orders belonging to their restaurant.
        if request.user.is_administrator():
            order = get_object_or_404(Order, id=order_id)
        else:
            session_restaurant_id = request.session.get('selected_restaurant_id')
            restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
            current_restaurant = restaurant_context['current_restaurant']
            view_all_restaurants = restaurant_context['view_all_restaurants']

            if current_restaurant and not view_all_restaurants:
                if current_restaurant.is_main_restaurant:
                    order = get_object_or_404(
                        Order,
                        Q(id=order_id) & (
                            Q(table_info__restaurant=current_restaurant) |
                            Q(table_info__owner=current_restaurant.main_owner, table_info__restaurant__isnull=True)
                        )
                    )
                else:
                    order = get_object_or_404(
                        Order,
                        Q(id=order_id) & (
                            Q(table_info__restaurant=current_restaurant) |
                            Q(table_info__owner=current_restaurant.branch_owner, table_info__restaurant__isnull=True)
                        )
                    )
            else:
                accessible_restaurants = restaurant_context['accessible_restaurants']
                if accessible_restaurants.exists():
                    order_query = Q()
                    for restaurant in accessible_restaurants:
                        order_query |= (
                            Q(table_info__restaurant=restaurant) |
                            Q(table_info__owner=restaurant.main_owner, table_info__restaurant__isnull=True) |
                            Q(table_info__owner=restaurant.branch_owner, table_info__restaurant__isnull=True)
                        )
                    order = get_object_or_404(Order, Q(id=order_id) & order_query)
                else:
                    return JsonResponse({'success': False, 'message': 'No restaurant context found.'})

        order_number = order.order_number
        _cleanup_order_dependencies([order.id])
        order.delete()

        return JsonResponse({
            'success': True,
            'message': f'Order #{order_number} deleted successfully'
        })

    except Exception as e:
        logger.error(f'Error deleting order: {str(e)}')
        return JsonResponse({'success': False, 'message': 'Error deleting order. Please try again.'})


@login_required
@require_POST
@transaction.atomic
def bulk_delete_orders(request):
    """Bulk delete multiple orders"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'error': 'Access denied. Administrator or Owner privileges required.'})
    
    try:
        # Parse JSON data
        data = json.loads(request.body)
        order_ids = data.get('order_ids', [])
        
        if not order_ids:
            return JsonResponse({'success': False, 'error': 'No orders selected for deletion.'})
        
        if len(order_ids) > 50:  # Limit bulk operations
            return JsonResponse({'success': False, 'error': 'Cannot delete more than 50 orders at once.'})
        
        # Get restaurant context for proper filtering
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        current_restaurant = restaurant_context['current_restaurant']
        view_all_restaurants = restaurant_context['view_all_restaurants']
        
        # Build query for orders to delete with proper restaurant filtering
        if current_restaurant and not view_all_restaurants:
            # Filter by specific selected restaurant
            if current_restaurant.is_main_restaurant:
                # Main restaurant: orders from tables assigned to it OR owned by main owner
                orders_to_delete = Order.objects.filter(
                    id__in=order_ids
                ).filter(
                    Q(table_info__restaurant=current_restaurant) |
                    Q(table_info__owner=current_restaurant.main_owner, table_info__restaurant__isnull=True)
                )
            else:
                # Branch restaurant: only orders from tables specifically assigned to this branch
                orders_to_delete = Order.objects.filter(
                    id__in=order_ids
                ).filter(
                    Q(table_info__restaurant=current_restaurant) |
                    Q(table_info__owner=current_restaurant.branch_owner, table_info__restaurant__isnull=True)
                )
        elif request.user.is_administrator():
            # Administrator sees all orders
            orders_to_delete = Order.objects.filter(id__in=order_ids)
        else:
            # Get orders from all accessible restaurants
            accessible_restaurants = restaurant_context['accessible_restaurants']
            
            if accessible_restaurants.exists():
                order_query = Q()
                for restaurant in accessible_restaurants:
                    order_query |= (
                        Q(table_info__restaurant=restaurant) |
                        Q(table_info__owner=restaurant.main_owner, table_info__restaurant__isnull=True) |
                        Q(table_info__owner=restaurant.branch_owner, table_info__restaurant__isnull=True)
                    )
                orders_to_delete = Order.objects.filter(id__in=order_ids).filter(order_query)
            else:
                orders_to_delete = Order.objects.none()
        
        # Check if all requested orders exist and are accessible
        found_count = orders_to_delete.count()
        if found_count != len(order_ids):
            return JsonResponse({
                'success': False, 
                'error': f'Some orders could not be found or you do not have permission to delete them. Found {found_count} out of {len(order_ids)} orders.'
            })
        
        # Get order numbers and IDs for logging and cleanup
        order_numbers = list(orders_to_delete.values_list('order_number', flat=True))
        order_ids = list(orders_to_delete.values_list('id', flat=True))

        # Clean up all PROTECT FK chains before bulk deletion
        _cleanup_order_dependencies(order_ids)

        # Perform bulk deletion
        deleted_count, deleted_details = orders_to_delete.delete()
        
        # Log the deletion
        if hasattr(request.user, 'get_full_name'):
            user_name = request.user.get_full_name() or request.user.username
        else:
            user_name = request.user.username
            
        logger.info(f"Bulk delete performed by {user_name}: {deleted_count} orders deleted - {', '.join(order_numbers[:5])}")
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} order(s).'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data.'})
    except Exception as e:
        logger.error(f"Error in bulk delete orders: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An error occurred while deleting orders.'})


@login_required
def profile(request):
    """User profile view - accessible by all admin panel users"""
    if not (request.user.is_administrator() or request.user.is_owner() or request.user.is_manager() or 
            request.user.is_kitchen_staff() or request.user.is_customer_care()):
        messages.error(request, "Access denied. Admin panel access required.")
        return redirect('restaurant:home')

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_profile':
            # Update profile information
            try:
                request.user.first_name = request.POST.get('first_name', '').strip()
                request.user.last_name = request.POST.get('last_name', '').strip()
                request.user.email = request.POST.get('email', '').strip()
                request.user.phone_number = request.POST.get('phone_number', '').strip()
                request.user.address = request.POST.get('address', '').strip()
                
                # Validation
                if not request.user.first_name or not request.user.last_name:
                    messages.error(request, "First name and last name are required.")
                    return redirect('admin_panel:profile')
                
                if request.user.email and User.objects.filter(email=request.user.email).exclude(id=request.user.id).exists():
                    messages.error(request, "Email already exists.")
                    return redirect('admin_panel:profile')
                
                request.user.save()
                messages.success(request, "Profile updated successfully.")
                
            except Exception as e:
                logger.error(f'Error updating profile: {str(e)}')
                messages.error(request, "Error updating profile. Please try again.")
                
        elif action == 'change_password':
            # Change password
            try:
                current_password = request.POST.get('current_password')
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')
                
                # Validation
                if not current_password or not new_password or not confirm_password:
                    messages.error(request, "All password fields are required.")
                    return redirect('admin_panel:profile')
                
                if not request.user.check_password(current_password):
                    messages.error(request, "Current password is incorrect.")
                    return redirect('admin_panel:profile')
                
                if new_password != confirm_password:
                    messages.error(request, "New passwords do not match.")
                    return redirect('admin_panel:profile')
                
                if len(new_password) < 8:
                    messages.error(request, "Password must be at least 8 characters long.")
                    return redirect('admin_panel:profile')
                
                request.user.set_password(new_password)
                request.user.save()
                update_session_auth_hash(request, request.user)
                messages.success(request, "Password changed successfully.")
                
            except Exception as e:
                logger.error(f'Error changing password: {str(e)}')
                messages.error(request, "Error changing password. Please try again.")
        
        elif action == 'update_restaurant':
            # Update restaurant information (for owners and managers)
            try:
                if not (request.user.is_owner() or request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
                    messages.error(request, "Only restaurant owners and managers can update restaurant information.")
                    return redirect('admin_panel:profile')
                
                restaurant_name = request.POST.get('restaurant_name', '').strip()
                restaurant_description = request.POST.get('restaurant_description', '').strip()
                tax_rate_percentage = request.POST.get('tax_rate_percentage', '').strip()
                
                logger.info(f"UPDATE_RESTAURANT: user={request.user.username}, tax_rate_percentage={tax_rate_percentage}")
                
                # Validation
                if not restaurant_name:
                    messages.error(request, "Restaurant name is required.")
                    return redirect('admin_panel:profile')
                
                # Validate tax rate
                try:
                    if tax_rate_percentage != '':  # Check explicitly for empty string (allow "0")
                        tax_rate_float = float(tax_rate_percentage)
                        if tax_rate_float < 0 or tax_rate_float > 99.99:
                            messages.error(request, "Tax rate must be between 0% and 99.99%.")
                            return redirect('admin_panel:profile')
                        
                        # Convert percentage to decimal for storage (0% = 0.00, 15% = 0.15, etc.)
                        tax_rate_decimal = Decimal(str(tax_rate_float / 100))
                    else:
                        # If field is empty, use default tax rate from User model default
                        tax_rate_decimal = Decimal('0.0800')  # Same as User model default
                        
                except (ValueError, InvalidOperation):
                    messages.error(request, "Invalid tax rate. Please enter a valid number.")
                    return redirect('admin_panel:profile')
                
                # Update restaurant information
                request.user.restaurant_name = restaurant_name
                request.user.restaurant_description = restaurant_description
                request.user.tax_rate = tax_rate_decimal
                
                logger.info(f"UPDATE_RESTAURANT: Setting user.tax_rate to {tax_rate_decimal}")
                
                # ALSO update Restaurant model tax_rate if user has a restaurant
                from restaurant.models_restaurant import Restaurant
                user_restaurants = Restaurant.objects.filter(
                    models.Q(main_owner=request.user) | models.Q(branch_owner=request.user)
                )
                logger.info(f"UPDATE_RESTAURANT: Found {user_restaurants.count()} restaurants for user")
                allow_remote_orders = request.POST.get('allow_remote_orders') == 'on'
                for restaurant in user_restaurants:
                    restaurant.tax_rate = tax_rate_decimal
                    restaurant.allow_remote_orders = allow_remote_orders
                    restaurant.save()
                    logger.info(f"UPDATE_RESTAURANT: Updated restaurant {restaurant.name} tax_rate to {tax_rate_decimal}")
                
                # Handle auto-print settings (checkboxes)
                request.user.auto_print_kot = request.POST.get('auto_print_kot') == 'on'
                request.user.auto_print_bot = request.POST.get('auto_print_bot') == 'on'
                
                # Handle WiFi settings
                wifi_ssid = request.POST.get('wifi_ssid', '').strip()
                wifi_password = request.POST.get('wifi_password', '').strip()
                
                # Update WiFi settings for all restaurants
                for restaurant in user_restaurants:
                    restaurant.wifi_ssid = wifi_ssid if wifi_ssid else None
                    restaurant.wifi_password = wifi_password if wifi_password else None
                    restaurant.save()
                
                # Handle logo upload
                if 'restaurant_logo' in request.FILES:
                    logo_file = request.FILES['restaurant_logo']
                    # Validate using Pillow (checks real file content, not just header)
                    from restaurant.models_restaurant import validate_restaurant_image
                    try:
                        validate_restaurant_image(logo_file)
                    except ValidationError as e:
                        messages.error(request, str(e.message))
                        return redirect('admin_panel:profile')
                    
                    # Update all restaurants for this user
                    for restaurant in user_restaurants:
                        if restaurant.logo:
                            restaurant.logo.delete()  # Delete old logo
                        restaurant.logo = logo_file
                        restaurant.save()
                
                # Generate QR code if it doesn't exist
                if not request.user.restaurant_qr_code:
                    request.user.generate_qr_code()
                
                request.user.save()
                logger.info(f"UPDATE_RESTAURANT: User saved. Final tax_rate={request.user.tax_rate}")
                messages.success(request, "Restaurant information updated successfully.")
                
            except Exception as e:
                logger.error(f'Error updating restaurant: {str(e)}')
                messages.error(request, "Error updating restaurant. Please try again.")
        
        elif action == 'update_currency':
            # Update currency settings (for owners and managers)
            try:
                if not (request.user.is_owner() or request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
                    messages.error(request, "Only restaurant owners and managers can update currency settings.")
                    return redirect('admin_panel:profile')
                
                currency_code = request.POST.get('currency_code', 'USD').strip().upper()
                
                # Validate currency code
                valid_currencies = ['USD', 'EUR', 'GBP', 'KES', 'TZS', 'UGX', 'RWF', 'ZAR', 'NGN', 'GHS', 'INR', 'AED', 'SAR', 'CNY', 'JPY']
                if currency_code not in valid_currencies:
                    messages.error(request, "Invalid currency code selected.")
                    return redirect('admin_panel:profile')
                
                # Update the user's currency code
                request.user.currency_code = currency_code
                request.user.save()
                
                # Also update the Restaurant model if it exists (for PRO plan branches)
                try:
                    from restaurant.models_restaurant import Restaurant
                    
                    # Update main restaurant if user is main_owner
                    if request.user.is_main_owner():
                        restaurants = Restaurant.objects.filter(main_owner=request.user)
                        restaurants.update(currency_code=currency_code)
                    # Update specific branch if user is branch_owner
                    elif request.user.is_branch_owner():
                        restaurants = Restaurant.objects.filter(branch_owner=request.user)
                        restaurants.update(currency_code=currency_code)
                    # For legacy owner role
                    else:
                        restaurants = Restaurant.objects.filter(
                            models.Q(main_owner=request.user) | models.Q(branch_owner=request.user)
                        )
                        restaurants.update(currency_code=currency_code)
                except Exception as e:
                    # Restaurant model update failed, but user currency is saved
                    pass
                
                messages.success(request, f"Currency updated to {currency_code} successfully.")
                
            except Exception as e:
                logger.error(f'Error updating currency: {str(e)}')
                messages.error(request, "Error updating currency. Please try again.")
        
        return redirect('admin_panel:profile')
    
    # Get restaurant context for all users (including managers)
    session_restaurant_id = request.session.get('selected_restaurant_id')
    restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
    
    context = {
        'user': request.user,
        **restaurant_context,  # Include restaurant context (restaurant_name, restaurant_logo, etc.)
    }
    
    # Get restaurant logo and WiFi settings if user is owner
    if request.user.is_owner():
        from restaurant.models_restaurant import Restaurant
        try:
            restaurant = Restaurant.objects.filter(
                models.Q(main_owner=request.user) | models.Q(branch_owner=request.user)
            ).first()
            if restaurant:
                if restaurant.logo:
                    context['restaurant_logo'] = restaurant.logo
                # Add WiFi settings to context
                context['wifi_ssid'] = restaurant.wifi_ssid or ''
                context['wifi_password'] = restaurant.wifi_password or ''
                context['current_restaurant'] = restaurant
        except:
            pass
    
    return render(request, 'admin_panel/profile.html', context)

@login_required
def manage_qr_code(request):
    """QR Code management for restaurant owners and branch owners"""
    if not (request.user.is_owner() or request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Owner or Branch Owner privileges required.")
        return redirect('restaurant:home')
    
    # Get current restaurant context
    current_restaurant = None
    selected_restaurant_id = request.session.get('selected_restaurant_id')
    
    if selected_restaurant_id:
        # selected_restaurant_id stores the User (owner) ID, not Restaurant ID
        try:
            selected_user = User.objects.get(id=selected_restaurant_id)
            
            # Get the Restaurant object for this user
            if selected_user.is_branch_owner():
                current_restaurant = Restaurant.objects.filter(branch_owner=selected_user, is_main_restaurant=False).first()
            else:
                current_restaurant = Restaurant.objects.filter(main_owner=selected_user, is_main_restaurant=True).first()
            
            # Verify access
            if current_restaurant and not current_restaurant.can_user_access(request.user):
                current_restaurant = None
        except User.DoesNotExist:
            current_restaurant = None
    
    # If no restaurant selected, get user's default restaurant
    if not current_restaurant:
        if request.user.is_branch_owner():
            current_restaurant = Restaurant.objects.filter(branch_owner=request.user).first()
        elif request.user.is_main_owner():
            current_restaurant = Restaurant.objects.filter(main_owner=request.user, is_main_restaurant=True).first()
        elif request.user.is_manager() and request.user.owner:
            # Managers access their owner's restaurant
            if request.user.owner.is_branch_owner():
                current_restaurant = Restaurant.objects.filter(branch_owner=request.user.owner).first()
            elif request.user.owner.is_main_owner():
                current_restaurant = Restaurant.objects.filter(main_owner=request.user.owner, is_main_restaurant=True).first()
            else:
                # Regular owner
                current_restaurant = Restaurant.objects.filter(
                    Q(main_owner=request.user.owner) | Q(branch_owner=request.user.owner)
                ).first()
        elif request.user.is_owner():
            # Legacy support
            user_restaurants = Restaurant.objects.filter(
                models.Q(main_owner=request.user) | models.Q(branch_owner=request.user)
            )
            current_restaurant = user_restaurants.first()
    
    if not current_restaurant:
        messages.error(request, "No restaurant found. Please contact administrator.")
        return redirect('admin_panel:admin_dashboard')
    
    # Auto-generate menu QR code if not exists
    if not current_restaurant.menu_qr_code:
        import uuid
        current_restaurant.menu_qr_code = f"MENU-{uuid.uuid4().hex[:12].upper()}"
        current_restaurant.save()
    
    # Generate the full QR URLs using helper function
    qr_url = get_production_qr_url(request, current_restaurant.qr_code)
    
    # Generate menu QR URL
    host = request.get_host()
    if 'localhost' in host or '127.0.0.1' in host:
        menu_qr_url = f'http://{host}/menu-view/{current_restaurant.menu_qr_code}/'
    else:
        menu_qr_url = f'https://hospitality.easyfixsoft.com/menu-view/{current_restaurant.menu_qr_code}/'
    
    context = {
        'user': request.user,
        'qr_code': current_restaurant.qr_code,
        'qr_url': qr_url,
        'menu_qr_code': current_restaurant.menu_qr_code,
        'menu_qr_url': menu_qr_url,
        'restaurant_name': current_restaurant.name,
        'current_restaurant': current_restaurant,
        'is_branch': not current_restaurant.is_main_restaurant,
        'restaurant_logo': current_restaurant.logo if current_restaurant.logo else None,
    }
    
    return render(request, 'admin_panel/manage_qr_code.html', context)

@login_required
@require_POST
def regenerate_qr_code(request):
    """Regenerate QR code for restaurant owner or branch owner"""
    if not (request.user.is_owner() or request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Owner or Branch Owner privileges required.")
        return redirect('restaurant:home')
    
    # Get current restaurant context
    current_restaurant = None
    selected_restaurant_id = request.session.get('selected_restaurant_id')
    
    if selected_restaurant_id:
        try:
            current_restaurant = Restaurant.objects.get(id=selected_restaurant_id)
            if not current_restaurant.can_user_access(request.user):
                current_restaurant = None
        except Restaurant.DoesNotExist:
            current_restaurant = None
    
    if not current_restaurant:
        if request.user.is_branch_owner():
            current_restaurant = Restaurant.objects.filter(branch_owner=request.user).first()
        elif request.user.is_main_owner():
            current_restaurant = Restaurant.objects.filter(main_owner=request.user, is_main_restaurant=True).first()
        elif request.user.is_manager() and request.user.owner:
            # Managers access their owner's restaurant
            if request.user.owner.is_branch_owner():
                current_restaurant = Restaurant.objects.filter(branch_owner=request.user.owner).first()
            elif request.user.owner.is_main_owner():
                current_restaurant = Restaurant.objects.filter(main_owner=request.user.owner, is_main_restaurant=True).first()
            else:
                # Regular owner
                current_restaurant = Restaurant.objects.filter(
                    Q(main_owner=request.user.owner) | Q(branch_owner=request.user.owner)
                ).first()
    
    if not current_restaurant:
        messages.error(request, "No restaurant found for QR code regeneration.")
        return redirect('admin_panel:admin_dashboard')
    
    # Generate new QR code
    import uuid
    current_restaurant.qr_code = f"REST-{uuid.uuid4().hex[:12].upper()}"
    current_restaurant.save()
    
    messages.success(request, f'QR code has been regenerated successfully for {current_restaurant.name}!')
    return redirect('admin_panel:manage_qr_code')

@login_required
def generate_qr_image(request):
    """Generate QR code image for restaurant owner or branch owner"""
    if not (request.user.is_owner() or request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return HttpResponse("Access denied", status=403)
    
    # Check if generating menu QR or ordering QR
    qr_type = request.GET.get('type', 'order')  # Default to ordering QR
    
    # Get current restaurant context
    current_restaurant = None
    selected_restaurant_id = request.session.get('selected_restaurant_id')
    
    if selected_restaurant_id:
        try:
            current_restaurant = Restaurant.objects.get(id=selected_restaurant_id)
            if not current_restaurant.can_user_access(request.user):
                current_restaurant = None
        except Restaurant.DoesNotExist:
            current_restaurant = None
    
    if not current_restaurant:
        if request.user.is_branch_owner():
            current_restaurant = Restaurant.objects.filter(branch_owner=request.user).first()
        elif request.user.is_main_owner():
            current_restaurant = Restaurant.objects.filter(main_owner=request.user, is_main_restaurant=True).first()
        elif request.user.is_manager() and request.user.owner:
            # Managers access their owner's restaurant
            if request.user.owner.is_branch_owner():
                current_restaurant = Restaurant.objects.filter(branch_owner=request.user.owner).first()
            elif request.user.owner.is_main_owner():
                current_restaurant = Restaurant.objects.filter(main_owner=request.user.owner, is_main_restaurant=True).first()
            else:
                # Regular owner
                current_restaurant = Restaurant.objects.filter(
                    Q(main_owner=request.user.owner) | Q(branch_owner=request.user.owner)
                ).first()
    
    if not current_restaurant:
        return HttpResponse("No restaurant found", status=404)
    
    # Generate appropriate QR code based on type
    if qr_type == 'menu':
        # Menu QR Code - view only
        if not current_restaurant.menu_qr_code:
            import uuid
            current_restaurant.menu_qr_code = f"MENU-{uuid.uuid4().hex[:12].upper()}"
            current_restaurant.save()
        
        # Generate menu QR URL
        host = request.get_host()
        if 'localhost' in host or '127.0.0.1' in host:
            qr_url = f'http://{host}/menu-view/{current_restaurant.menu_qr_code}/'
        else:
            qr_url = f'https://hospitality.easyfixsoft.com/menu-view/{current_restaurant.menu_qr_code}/'
    else:
        # Ordering QR Code - default
        if not current_restaurant.qr_code:
            import uuid
            current_restaurant.qr_code = f"REST-{uuid.uuid4().hex[:12].upper()}"
            current_restaurant.save()
        
        # Generate ordering QR URL
        qr_url = get_production_qr_url(request, current_restaurant.qr_code)
    
    # Create QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(qr_url)
    qr.make(fit=True)

    # Create QR code image
    qr_img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to BytesIO
    img_io = io.BytesIO()
    qr_img.save(img_io, format='PNG')
    img_io.seek(0)
    
    # Return image response with NO CACHING to prevent stale QR codes
    response = HttpResponse(img_io.getvalue(), content_type='image/png')
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required
def debug_qr_code(request):
    """Debug endpoint to show QR code information - restricted to administrators only"""
    from django.conf import settings
    from django.http import JsonResponse

    # Only system administrators may access this diagnostic endpoint, regardless of DEBUG mode
    if not request.user.is_administrator():
        logger.warning(
            f"Unauthorized debug_qr_code access attempt by {request.user.username} "
            f"from {request.META.get('REMOTE_ADDR')}"
        )
        return HttpResponse("This endpoint is not available.", status=403)

    qr_url = get_production_qr_url(request, request.user.restaurant_qr_code)

    debug_info = {
        'username': request.user.username,
        'restaurant_name': request.user.restaurant_name,
        'qr_code_in_database': request.user.restaurant_qr_code,
        'full_qr_url': qr_url,
        'expected_access_url': f"/r/{request.user.restaurant_qr_code}/",
        'is_owner': request.user.is_owner(),
        'host': request.get_host(),
        'debug_mode': settings.DEBUG,
    }

    return JsonResponse(debug_info, json_dumps_params={'indent': 2})

@login_required
def import_products_csv(request):
    """Import products from CSV file"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    if request.method != 'POST':
        return redirect('admin_panel:manage_products')
    
    if 'csv_file' not in request.FILES:
        messages.error(request, 'Please select a CSV file to upload.')
        return redirect('admin_panel:manage_products')
    
    csv_file = request.FILES['csv_file']
    
    # Validate file type
    if not csv_file.name.endswith('.csv'):
        messages.error(request, 'Please upload a valid CSV file.')
        return redirect('admin_panel:manage_products')
    
    try:
        # Get restaurant context
        from restaurant.models_restaurant import Restaurant
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get restaurant from form or use current restaurant
        restaurant_id = request.POST.get('restaurant')
        if restaurant_id:
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
                # Verify user has access to this restaurant
                accessible_restaurants = restaurant_context.get('accessible_restaurants')
                if accessible_restaurants is None or not accessible_restaurants.filter(id=target_restaurant.id).exists():
                    messages.error(request, 'You do not have permission to import products to this restaurant')
                    return redirect('admin_panel:manage_products')
                current_restaurant = target_restaurant
            except Restaurant.DoesNotExist:
                messages.error(request, 'Selected restaurant does not exist')
                return redirect('admin_panel:manage_products')
        else:
            current_restaurant = restaurant_context.get('current_restaurant')
            if not current_restaurant:
                messages.error(request, 'Please select a restaurant for import')
                return redirect('admin_panel:manage_products')
        
        owner_filter = get_owner_filter(request.user)
        
        # Read CSV file - try multiple encodings
        file_content = csv_file.read()
        decoded_file = None
        
        # Try different encodings
        encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        for encoding in encodings:
            try:
                decoded_file = file_content.decode(encoding)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        
        if decoded_file is None:
            messages.error(request, 'Could not read the CSV file. Please save it with UTF-8 encoding.')
            return redirect('admin_panel:manage_products')
        
        csv_data = csv.DictReader(io.StringIO(decoded_file))
        
        imported_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        
        for row_num, row in enumerate(csv_data, start=2):  # Start at 2 because row 1 is header
            try:
                # Validate required fields
                name = row.get('name', '').strip()
                price = row.get('price', '').strip()
                main_category_name = row.get('main_category', '').strip()
                
                if not name:
                    errors.append(f"Row {row_num}: Product name is required")
                    error_count += 1
                    continue
                
                if not price:
                    errors.append(f"Row {row_num}: Price is required")
                    error_count += 1
                    continue
                
                if not main_category_name:
                    errors.append(f"Row {row_num}: Main category is required")
                    error_count += 1
                    continue
                
                # Validate price
                try:
                    price_decimal = Decimal(str(price))
                    if price_decimal <= 0:
                        errors.append(f"Row {row_num}: Price must be greater than 0")
                        error_count += 1
                        continue
                except (InvalidOperation, ValueError):
                    errors.append(f"Row {row_num}: Invalid price format")
                    error_count += 1
                    continue
                
                # Find or create main category for the selected restaurant
                main_category_filter = {'name__iexact': main_category_name, 'restaurant': current_restaurant}
                
                main_category = MainCategory.objects.filter(**main_category_filter).first()
                if not main_category:
                    # Determine owner based on restaurant type
                    if current_restaurant.is_main_restaurant:
                        category_owner = current_restaurant.main_owner
                    else:
                        category_owner = current_restaurant.branch_owner or current_restaurant.main_owner
                    
                    main_category_data = {
                        'name': main_category_name,
                        'is_active': True,
                        'description': row.get('main_category_description', '').strip(),
                        'owner': category_owner,
                        'restaurant': current_restaurant
                    }
                    main_category = MainCategory.objects.create(**main_category_data)
                    messages.info(request, f"Created main category '{main_category_name}' for this import.")
                
                # Handle subcategory
                sub_category = None
                sub_category_name = row.get('sub_category', '').strip()
                if sub_category_name:
                    sub_category = SubCategory.objects.filter(
                        name__iexact=sub_category_name,
                        main_category=main_category
                    ).first()
                    if not sub_category:
                        sub_category = SubCategory.objects.create(
                            main_category=main_category,
                            name=sub_category_name,
                            description=row.get('sub_category_description', '').strip(),
                            is_active=True
                        )
                        messages.info(request, f"Created sub category '{sub_category_name}' under '{main_category_name}'.")
                
                # Get station (default to kitchen if not specified)
                station = row.get('station', '').strip().lower()
                if station not in ['kitchen', 'bar', 'buffet', 'service']:
                    station = 'kitchen'  # Default to kitchen
                
                # Prepare product data
                product_data = {
                    'name': name,
                    'description': row.get('description', '').strip(),
                    'main_category': main_category,
                    'sub_category': sub_category,
                    'price': price_decimal,
                    'available_in_stock': max(0, int(row.get('available_in_stock', 0) or 0)),
                    'is_available': str(row.get('is_available', 'true')).lower() in ['true', '1', 'yes', 'available'],
                    'preparation_time': max(1, int(row.get('preparation_time', 15) or 15)),
                    'station': station,
                }
                
                # Check if product already exists - UPDATE if exists, CREATE if not
                existing_product = Product.objects.filter(
                    name__iexact=name,
                    main_category=main_category
                ).first()
                
                if existing_product:
                    # Update existing product
                    for field, value in product_data.items():
                        setattr(existing_product, field, value)
                    existing_product.save()
                    updated_count += 1
                else:
                    # Create new product
                    Product.objects.create(**product_data)
                imported_count += 1
                
            except Exception as e:
                logger.warning(f'CSV import row {row_num} error: {str(e)}')
                errors.append(f"Row {row_num}: Invalid data — check all fields and try again.")
                error_count += 1
                continue
        
        # Show results
        success_messages = []
        if imported_count > 0:
            success_messages.append(f'Created {imported_count} new products')
        if updated_count > 0:
            success_messages.append(f'Updated {updated_count} existing products')
        
        if success_messages:
            messages.success(request, f'Import completed! {", ".join(success_messages)}.')
        
        if error_count > 0:
            error_message = f'{error_count} errors occurred during import:'
            if len(errors) <= 10:
                error_message += '\n' + '\n'.join(errors)
            else:
                error_message += '\n' + '\n'.join(errors[:10]) + f'\n... and {len(errors) - 10} more errors'
            messages.error(request, error_message)
        
        if imported_count == 0 and updated_count == 0 and error_count == 0:
            messages.warning(request, 'No data found in the CSV file.')
            
    except Exception as e:
        logger.error(f'Error processing CSV file: {str(e)}')
        messages.error(request, 'Error processing CSV file. Please check the format and try again.')
    
    return redirect('admin_panel:manage_products')


@login_required
def import_products_excel(request):
    """Import products from Excel file"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    if request.method != 'POST':
        return redirect('admin_panel:manage_products')
    
    if 'excel_file' not in request.FILES:
        messages.error(request, 'Please select an Excel file to upload.')
        return redirect('admin_panel:manage_products')
    
    excel_file = request.FILES['excel_file']
    
    # Validate file type
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls).')
        return redirect('admin_panel:manage_products')
    
    if not openpyxl:
        messages.error(request, 'Excel import is not available. Please contact administrator.')
        return redirect('admin_panel:manage_products')
    
    try:
        # Get restaurant context
        from restaurant.models_restaurant import Restaurant
        session_restaurant_id = request.session.get('selected_restaurant_id')
        restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
        
        # Get restaurant from form or use current restaurant
        restaurant_id = request.POST.get('restaurant')
        if restaurant_id:
            try:
                target_restaurant = Restaurant.objects.get(id=restaurant_id)
                # Verify user has access to this restaurant
                accessible_restaurants = restaurant_context.get('accessible_restaurants')
                if accessible_restaurants is None or not accessible_restaurants.filter(id=target_restaurant.id).exists():
                    messages.error(request, 'You do not have permission to import products to this restaurant')
                    return redirect('admin_panel:manage_products')
                current_restaurant = target_restaurant
            except Restaurant.DoesNotExist:
                messages.error(request, 'Selected restaurant does not exist')
                return redirect('admin_panel:manage_products')
        else:
            current_restaurant = restaurant_context.get('current_restaurant')
            if not current_restaurant:
                messages.error(request, 'Please select a restaurant for import')
                return redirect('admin_panel:manage_products')
        
        owner_filter = get_owner_filter(request.user)
        
        # Save file temporarily
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        for chunk in excel_file.chunks():
            temp_file.write(chunk)
        temp_file.close()
        
        try:
            # Read Excel file
            workbook = openpyxl.load_workbook(temp_file.name, read_only=True, data_only=True)
            
            # Use first sheet or 'Products' sheet if exists
            sheet_name = 'Products' if 'Products' in workbook.sheetnames else workbook.sheetnames[0]
            worksheet = workbook[sheet_name]
            
            # Get header row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value).lower().strip())
                else:
                    headers.append('')
            
            # Map column indices
            col_mapping = {}
            for i, header in enumerate(headers):
                if 'name' in header and 'name' not in col_mapping:
                    col_mapping['name'] = i
                elif 'price' in header:
                    col_mapping['price'] = i
                elif 'category' in header and 'sub' not in header:
                    col_mapping['main_category'] = i
                elif 'sub' in header and 'category' in header:
                    col_mapping['sub_category'] = i
                elif 'description' in header:
                    col_mapping['description'] = i
                elif 'stock' in header or 'quantity' in header:
                    col_mapping['available_in_stock'] = i
                elif 'available' in header or 'status' in header:
                    col_mapping['is_available'] = i
                elif 'time' in header and 'prep' in header:
                    col_mapping['preparation_time'] = i
                elif 'station' in header:
                    col_mapping['station'] = i
            
            if 'name' not in col_mapping or 'price' not in col_mapping or 'main_category' not in col_mapping:
                messages.error(request, 'Excel file must contain columns for Name, Price, and Main Category.')
                return redirect('admin_panel:manage_products')
            
            imported_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            
            # Process data rows
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):  # Skip empty rows
                        continue
                    
                    # Extract data
                    name = str(row[col_mapping['name']] or '').strip()
                    price = str(row[col_mapping['price']] or '').strip()
                    main_category_name = str(row[col_mapping['main_category']] or '').strip()
                    
                    if not name:
                        errors.append(f"Row {row_num}: Product name is required")
                        error_count += 1
                        continue
                    
                    if not price:
                        errors.append(f"Row {row_num}: Price is required")
                        error_count += 1
                        continue
                    
                    if not main_category_name:
                        errors.append(f"Row {row_num}: Main category is required")
                        error_count += 1
                        continue
                    
                    # Validate price
                    try:
                        price_decimal = Decimal(str(price))
                        if price_decimal <= 0:
                            errors.append(f"Row {row_num}: Price must be greater than 0")
                            error_count += 1
                            continue
                    except (InvalidOperation, ValueError):
                        errors.append(f"Row {row_num}: Invalid price format")
                        error_count += 1
                        continue
                    
                    # Find or CREATE main category for the selected restaurant
                    main_category_filter = {'name__iexact': main_category_name, 'restaurant': current_restaurant}
                    
                    main_category = MainCategory.objects.filter(**main_category_filter).first()
                    if not main_category:
                        # Auto-create main category (same logic as CSV import)
                        if current_restaurant.is_main_restaurant:
                            category_owner = current_restaurant.main_owner
                        else:
                            category_owner = current_restaurant.branch_owner or current_restaurant.main_owner
                        
                        main_category = MainCategory.objects.create(
                            name=main_category_name,
                            is_active=True,
                            description='',
                            owner=category_owner,
                            restaurant=current_restaurant
                        )
                        messages.info(request, f"Created main category '{main_category_name}' for this import.")
                    
                    # Handle subcategory - CREATE if not exists
                    sub_category = None
                    if 'sub_category' in col_mapping:
                        sub_category_name = str(row[col_mapping['sub_category']] or '').strip()
                        if sub_category_name:
                            sub_category = SubCategory.objects.filter(
                                name__iexact=sub_category_name,
                                main_category=main_category
                            ).first()
                            if not sub_category:
                                # Auto-create subcategory
                                sub_category = SubCategory.objects.create(
                                    main_category=main_category,
                                    name=sub_category_name,
                                    description='',
                                    is_active=True
                                )
                                messages.info(request, f"Created sub category '{sub_category_name}' under '{main_category_name}'.")
                    
                    # Prepare product data
                    product_data = {
                        'name': name,
                        'description': str(row[col_mapping.get('description', 0)] or '').strip(),
                        'main_category': main_category,
                        'sub_category': sub_category,
                        'price': price_decimal,
                        'available_in_stock': max(0, int(row[col_mapping.get('available_in_stock', 0)] or 0)),
                        'preparation_time': max(1, int(row[col_mapping.get('preparation_time', 0)] or 15)),
                    }
                    
                    # Handle availability
                    if 'is_available' in col_mapping:
                        available_val = str(row[col_mapping['is_available']] or 'true').lower()
                        product_data['is_available'] = available_val in ['true', '1', 'yes', 'available']
                    else:
                        product_data['is_available'] = True
                    
                    # Handle station (default to kitchen if not specified)
                    if 'station' in col_mapping:
                        station = str(row[col_mapping['station']] or '').strip().lower()
                        product_data['station'] = station if station in ['kitchen', 'bar', 'buffet', 'service'] else 'kitchen'
                    else:
                        product_data['station'] = 'kitchen'
                    
                    # Check if product exists - UPDATE if exists, CREATE if not
                    existing_product = Product.objects.filter(
                        name__iexact=name,
                        main_category=main_category
                    ).first()
                    
                    if existing_product:
                        # Update existing product
                        for field, value in product_data.items():
                            setattr(existing_product, field, value)
                        existing_product.save()
                        updated_count += 1
                    else:
                        # Create new product
                        Product.objects.create(**product_data)
                    imported_count += 1
                    
                except Exception as e:
                    logger.warning(f'Excel import row {row_num} error: {str(e)}')
                    errors.append(f"Row {row_num}: Invalid data — check all fields and try again.")
                    error_count += 1
                    continue
            
            workbook.close()
            
            # Show results
            success_messages = []
            if imported_count > 0:
                success_messages.append(f'Created {imported_count} new products')
            if updated_count > 0:
                success_messages.append(f'Updated {updated_count} existing products')
            
            if success_messages:
                messages.success(request, f'Import completed! {", ".join(success_messages)}.')
            
            if error_count > 0:
                error_message = f'{error_count} errors occurred during import:'
                if len(errors) <= 10:
                    error_message += '\n' + '\n'.join(errors)
                else:
                    error_message += '\n' + '\n'.join(errors[:10]) + f'\n... and {len(errors) - 10} more errors'
                messages.error(request, error_message)
            
            if imported_count == 0 and updated_count == 0 and error_count == 0:
                messages.warning(request, 'No data found in the Excel file.')
                
        finally:
            # Clean up temp file
            if os.path.exists(temp_file.name):
                os.unlink(temp_file.name)
                
    except Exception as e:
        logger.error(f'Error processing Excel file: {str(e)}')
        messages.error(request, 'Error processing Excel file. Please check the format and try again.')
    
    return redirect('admin_panel:manage_products')


@login_required
def download_template_csv(request):
    """Download CSV template for product import"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="product_import_template.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'name',
        'description',
        'main_category',
        'sub_category',
        'price',
        'available_in_stock',
        'is_available',
        'preparation_time',
        'station'
    ])
    
    # Add sample data
    writer.writerow([
        'Sample Pizza',
        'Delicious cheese pizza with fresh toppings',
        'Main Dishes',
        'Pizza',
        '12.99',
        '50',
        'true',
        '20',
        'kitchen'
    ])
    writer.writerow([
        'Sample Burger',
        'Juicy beef burger with lettuce and tomato',
        'Main Dishes',
        'Burgers',
        '8.99',
        '30',
        'true',
        '15',
        'kitchen'
    ])
    writer.writerow([
        'Sample Cocktail',
        'Refreshing tropical cocktail',
        'Beverages',
        'Alcoholic',
        '7.50',
        '20',
        'true',
        '5',
        'bar'
    ])
    
    return response


@login_required
def download_template_excel(request):
    """Download Excel template for product import"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    if not openpyxl:
        messages.error(request, 'Excel export is not available. Please contact administrator.')
        return redirect('admin_panel:manage_products')
    
    # Create workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"
    
    # Headers
    headers = [
        'name',
        'description', 
        'main_category',
        'sub_category',
        'price',
        'available_in_stock',
        'is_available',
        'preparation_time',
        'station'
    ]
    
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # Sample data
    sample_data = [
        ['Sample Pizza', 'Delicious cheese pizza with fresh toppings', 'Main Dishes', 'Pizza', 12.99, 50, True, 20, 'kitchen'],
        ['Sample Burger', 'Juicy beef burger with lettuce and tomato', 'Main Dishes', 'Burgers', 8.99, 30, True, 15, 'kitchen'],
        ['Sample Cocktail', 'Refreshing tropical cocktail', 'Beverages', 'Alcoholic', 7.50, 20, True, 5, 'bar'],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to BytesIO
    excel_io = io.BytesIO()
    workbook.save(excel_io)
    excel_io.seek(0)
    
    response = HttpResponse(
        excel_io.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="product_import_template.xlsx"'
    
    return response


@login_required
@require_POST
@transaction.atomic
def bulk_delete_products(request):
    """Bulk delete multiple products"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'error': 'Access denied. Administrator or Owner privileges required.'})
    
    try:
        # Parse JSON data
        data = json.loads(request.body)
        product_ids = data.get('product_ids', [])
        
        if not product_ids:
            return JsonResponse({'success': False, 'error': 'No products selected for deletion.'})
        
        if len(product_ids) > 50:  # Limit bulk operations
            return JsonResponse({'success': False, 'error': 'Cannot delete more than 50 products at once.'})
        
        owner_filter = get_owner_filter(request.user)
        
        # Build query for products to delete
        if owner_filter:
            # Owner can only delete their own products
            products_to_delete = Product.objects.filter(
                id__in=product_ids
            ).filter(
                Q(main_category__owner=owner_filter) |
                Q(main_category__restaurant__main_owner=owner_filter) |
                Q(main_category__restaurant__branch_owner=owner_filter)
            )
        else:
            # Administrator can delete all products
            products_to_delete = Product.objects.filter(id__in=product_ids)
        
        # Check if all requested products exist and are accessible
        found_count = products_to_delete.count()
        if found_count != len(product_ids):
            return JsonResponse({
                'success': False, 
                'error': f'Some products could not be found or you do not have permission to delete them. Found {found_count} out of {len(product_ids)} products.'
            })
        
        # Check for products that are in active orders (optional business logic)
        # You can uncomment this if you want to prevent deletion of products with active orders
        # from orders.models import OrderItem
        # active_order_products = OrderItem.objects.filter(
        #     product__in=products_to_delete,
        #     order__status__in=['pending', 'confirmed', 'preparing']
        # ).values_list('product_id', flat=True).distinct()
        # 
        # if active_order_products:
        #     return JsonResponse({
        #         'success': False,
        #         'error': f'Cannot delete products that are in active orders. {len(active_order_products)} products have active orders.'
        #     })
        
        # Get product names and IDs for logging and cleanup
        product_names = list(products_to_delete.values_list('name', flat=True))
        product_ids = list(products_to_delete.values_list('id', flat=True))

        # Nullify all PROTECT FK refs before bulk deletion
        _cleanup_product_dependencies(product_ids)

        # Perform bulk deletion
        deleted_count, deleted_details = products_to_delete.delete()
        
        # Log the deletion
        if hasattr(request.user, 'get_full_name'):
            user_name = request.user.get_full_name() or request.user.username
        else:
            user_name = request.user.username
            
        logger.info(f"Bulk delete performed by {user_name}: {deleted_count} products deleted - {', '.join(product_names[:5])}")
        
        return JsonResponse({
            'success': True,
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} products.'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data.'})
    except Exception as e:
        logger.error(f"Error in bulk delete: {str(e)}")
        return JsonResponse({'success': False, 'error': 'An error occurred while deleting products.'})


@login_required
def export_products_csv(request):
    """Export all products to CSV"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    # Get restaurant context for proper filtering
    session_restaurant_id = request.session.get('selected_restaurant_id')
    restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
    current_restaurant = restaurant_context['current_restaurant']
    view_all_restaurants = restaurant_context['view_all_restaurants']
    accessible_restaurants = restaurant_context['accessible_restaurants']
    
    # Get products based on restaurant context
    if view_all_restaurants:
        # Export from all accessible restaurants
        products = Product.objects.filter(main_category__restaurant__in=accessible_restaurants)
    elif current_restaurant:
        # Export from current restaurant only
        products = Product.objects.filter(main_category__restaurant=current_restaurant)
    else:
        products = Product.objects.none()
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'name', 'description', 'main_category', 'sub_category', 
        'price', 'available_in_stock', 'is_available', 
        'preparation_time', 'station'
    ])
    
    for product in products:
        writer.writerow([
            _safe_csv(product.name),
            _safe_csv(product.description),
            _safe_csv(product.main_category.name),
            _safe_csv(product.sub_category.name) if product.sub_category else '',
            product.price,
            product.available_in_stock,
            product.is_available,
            product.preparation_time,
            product.station
        ])
    
    return response


@login_required  
def export_products_excel(request):
    """Export all products to Excel"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    try:
        import openpyxl
    except ImportError:
        messages.error(request, 'Excel export is not available. Please contact administrator.')
        return redirect('admin_panel:manage_products')
    
    # Get restaurant context for proper filtering
    session_restaurant_id = request.session.get('selected_restaurant_id')
    restaurant_context = get_restaurant_context(request.user, session_restaurant_id, request)
    current_restaurant = restaurant_context['current_restaurant']
    view_all_restaurants = restaurant_context['view_all_restaurants']
    accessible_restaurants = restaurant_context['accessible_restaurants']
    
    # Get products based on restaurant context
    if view_all_restaurants:
        # Export from all accessible restaurants
        products = Product.objects.filter(main_category__restaurant__in=accessible_restaurants)
    elif current_restaurant:
        # Export from current restaurant only
        products = Product.objects.filter(main_category__restaurant=current_restaurant)
    else:
        products = Product.objects.none()
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Products Export"
    
    # Headers
    headers = [
        'Name', 'Description', 'Main Category', 'Sub Category',
        'Price', 'Available Stock', 'Available', 'Preparation Time', 'Station'
    ]
    
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # Data
    for row, product in enumerate(products, 2):
        worksheet.cell(row=row, column=1, value=product.name)
        worksheet.cell(row=row, column=2, value=product.description)
        worksheet.cell(row=row, column=3, value=product.main_category.name)
        worksheet.cell(row=row, column=4, value=product.sub_category.name if product.sub_category else '')
        worksheet.cell(row=row, column=5, value=float(product.price))
        worksheet.cell(row=row, column=6, value=product.available_in_stock)
        worksheet.cell(row=row, column=7, value=product.is_available)
        worksheet.cell(row=row, column=8, value=product.preparation_time)
        worksheet.cell(row=row, column=9, value=product.station)
    
    # Save to BytesIO
    excel_io = io.BytesIO()
    workbook.save(excel_io)
    excel_io.seek(0)
    
    response = HttpResponse(
        excel_io.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'
    return response


@login_required
def export_products_pdf(request):
    """Export all products to PDF"""
    if not (request.user.is_administrator() or request.user.is_owner() or 
            request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, "Access denied. Administrator or Owner privileges required.")
        return redirect('admin_panel:manage_products')
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        messages.error(request, 'PDF export is not available. Please contact administrator.')
        return redirect('admin_panel:manage_products')
    
    # Get user's products
    if request.user.is_administrator():
        products = Product.objects.all()
        restaurant_name = "All Restaurants"
    else:
        products = Product.objects.filter(
            Q(main_category__owner=request.user) |
            Q(main_category__restaurant__main_owner=request.user) |
            Q(main_category__restaurant__branch_owner=request.user)
        ).distinct()
        restaurant_name = request.user.restaurant_name or "Restaurant"
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="products_export.pdf"'
    
    # Create PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    
    elements = []
    elements.append(Paragraph(f"{restaurant_name} - Products List", title_style))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['Name', 'Category', 'Price', 'Stock', 'Station', 'Available']]
    
    for product in products:
        data.append([
            product.name[:30],  # Truncate long names
            product.main_category.name[:20],
            f"${product.price}",
            str(product.available_in_stock),
            product.station.title(),
            "Yes" if product.is_available else "No"
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response


# ============================================================================
# Printer Configuration Views
# ============================================================================

@login_required
def printer_settings(request):
    """Display printer configuration page"""
    if not (request.user.is_owner() or request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        messages.error(request, 'Only restaurant owners can access printer settings.')
        return redirect('admin_panel:admin_dashboard')
    
    user = request.user
    
    # Get the current restaurant based on session or user type (including managers via their owner)
    from restaurant.models_restaurant import Restaurant
    
    # Use restaurant context helper to properly handle managers and all owner types
    session_restaurant_id = request.session.get('selected_restaurant_id')
    restaurant_context_data = get_restaurant_context(user, session_restaurant_id, request)
    current_restaurant = restaurant_context_data['current_restaurant']
    
    # Get printer settings - prefer Restaurant model, fallback to User model
    if current_restaurant:
        kitchen_printer = current_restaurant.kitchen_printer_name or ''
        bar_printer = current_restaurant.bar_printer_name or ''
        buffet_printer = current_restaurant.buffet_printer_name or ''
        service_printer = current_restaurant.service_printer_name or ''
        receipt_printer = current_restaurant.receipt_printer_name or ''
        auto_print_kot = current_restaurant.auto_print_kot
        auto_print_bot = current_restaurant.auto_print_bot
        auto_print_buffet = current_restaurant.auto_print_buffet
        auto_print_service = current_restaurant.auto_print_service
        settings_source = 'restaurant'
    else:
        kitchen_printer = user.kitchen_printer_name or ''
        bar_printer = user.bar_printer_name or ''
        buffet_printer = getattr(user, 'buffet_printer_name', '') or ''
        service_printer = getattr(user, 'service_printer_name', '') or ''
        receipt_printer = user.receipt_printer_name or ''
        auto_print_kot = user.auto_print_kot
        auto_print_bot = user.auto_print_bot
        auto_print_buffet = getattr(user, 'auto_print_buffet', True)
        auto_print_service = getattr(user, 'auto_print_service', True)
        settings_source = 'user'
    
    context = {
        'owner': user,
        'current_restaurant': current_restaurant,
        'restaurant_name': current_restaurant.name if current_restaurant else 'Restaurant System',
        'settings_source': settings_source,
        'kitchen_printer': kitchen_printer,
        'bar_printer': bar_printer,
        'buffet_printer': buffet_printer,
        'service_printer': service_printer,
        'receipt_printer': receipt_printer,
        'auto_print_kot': auto_print_kot,
        'auto_print_bot': auto_print_bot,
        'auto_print_buffet': auto_print_buffet,
        'auto_print_service': auto_print_service,
        'restaurant_logo': current_restaurant.logo if current_restaurant and current_restaurant.logo else None,
    }
    
    # Get or create API token for print client
    try:
        from rest_framework.authtoken.models import Token
        token, created = Token.objects.get_or_create(user=user)
        context['api_token'] = token.key
    except Exception:
        context['api_token'] = None
    
    return render(request, 'admin_panel/printer_settings.html', context)


@login_required
@require_POST
def save_printer_settings(request):
    """Save printer configuration - saves to Restaurant model for proper per-restaurant settings"""
    # Return JSON for AJAX requests even if permission denied
    if not (request.user.is_owner() or request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        user = request.user
        from restaurant.models_restaurant import Restaurant
        
        # Get form data - handle both FormData and JSON
        if request.content_type and 'application/json' in request.content_type:
            import json
            data = json.loads(request.body)
            kitchen_printer = data.get('kitchen_printer', '').strip()
            bar_printer = data.get('bar_printer', '').strip()
            buffet_printer = data.get('buffet_printer', '').strip()
            service_printer = data.get('service_printer', '').strip()
            receipt_printer = data.get('receipt_printer', '').strip()
            auto_print_kot = data.get('auto_print_kot', False)
            auto_print_bot = data.get('auto_print_bot', False)
            auto_print_buffet = data.get('auto_print_buffet', False)
            auto_print_service = data.get('auto_print_service', False)
        else:
            kitchen_printer = request.POST.get('kitchen_printer', '').strip()
            bar_printer = request.POST.get('bar_printer', '').strip()
            buffet_printer = request.POST.get('buffet_printer', '').strip()
            service_printer = request.POST.get('service_printer', '').strip()
            receipt_printer = request.POST.get('receipt_printer', '').strip()
            auto_print_kot = request.POST.get('auto_print_kot') == 'on'
            auto_print_bot = request.POST.get('auto_print_bot') == 'on'
            auto_print_buffet = request.POST.get('auto_print_buffet') == 'on'
            auto_print_service = request.POST.get('auto_print_service') == 'on'
        
        # Get the current restaurant based on session or user type
        current_restaurant = None
        session_restaurant_id = request.session.get('selected_restaurant_id')
        
        if session_restaurant_id:
            try:
                current_restaurant = Restaurant.objects.get(id=session_restaurant_id)
            except Restaurant.DoesNotExist:
                pass
        
        # If no session restaurant, get the user's primary restaurant
        if not current_restaurant:
            if user.is_main_owner():
                current_restaurant = Restaurant.objects.filter(main_owner=user, is_main_restaurant=True).first()
            elif user.is_branch_owner():
                current_restaurant = Restaurant.objects.filter(branch_owner=user, is_main_restaurant=False).first()
            elif user.is_manager() and user.owner:
                # Managers: get their owner's restaurant
                if user.owner.is_main_owner():
                    current_restaurant = Restaurant.objects.filter(main_owner=user.owner, is_main_restaurant=True).first()
                elif user.owner.is_branch_owner():
                    current_restaurant = Restaurant.objects.filter(branch_owner=user.owner).first()
                else:
                    current_restaurant = Restaurant.objects.filter(main_owner=user.owner).first()
            elif user.is_owner():
                current_restaurant = Restaurant.objects.filter(main_owner=user).first()
        
        # ALWAYS save to User model first (ensures backup)
        user.kitchen_printer_name = kitchen_printer if kitchen_printer else None
        user.bar_printer_name = bar_printer if bar_printer else None
        if hasattr(user, 'buffet_printer_name'):
            user.buffet_printer_name = buffet_printer if buffet_printer else None
        if hasattr(user, 'service_printer_name'):
            user.service_printer_name = service_printer if service_printer else None
        user.receipt_printer_name = receipt_printer if receipt_printer else None
        user.auto_print_kot = auto_print_kot
        user.auto_print_bot = auto_print_bot
        if hasattr(user, 'auto_print_buffet'):
            user.auto_print_buffet = auto_print_buffet
        if hasattr(user, 'auto_print_service'):
            user.auto_print_service = auto_print_service
        user.save()
        
        # ALSO save to Restaurant model if available (for branch support)
        if current_restaurant:
            current_restaurant.kitchen_printer_name = kitchen_printer if kitchen_printer else None
            current_restaurant.bar_printer_name = bar_printer if bar_printer else None
            current_restaurant.buffet_printer_name = buffet_printer if buffet_printer else None
            current_restaurant.service_printer_name = service_printer if service_printer else None
            current_restaurant.receipt_printer_name = receipt_printer if receipt_printer else None
            current_restaurant.auto_print_kot = auto_print_kot
            current_restaurant.auto_print_bot = auto_print_bot
            current_restaurant.auto_print_buffet = auto_print_buffet
            current_restaurant.auto_print_service = auto_print_service
            current_restaurant.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Printer settings saved for {current_restaurant.name}!'
            })
        else:
            return JsonResponse({
                'success': True,
                'message': 'Printer settings saved successfully!'
            })
        
    except Exception as e:
        logger.exception(f"ERROR in save_printer_settings: {e}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to save printer settings. Please try again.'
        }, status=400)


@login_required
def detect_printers(request):
    """Detect available printers on the system"""
    try:
        import win32print  # type: ignore
        
        printers = []
        default_printer = None
        
        try:
            default_printer = win32print.GetDefaultPrinter()
        except Exception:
            pass
        
        # Get all local printers
        try:
            for printer_info in win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL):
                printer_name = printer_info[2]
                printers.append({
                    'name': printer_name,
                    'is_default': printer_name == default_printer
                })
        except Exception as enum_error:
            # If EnumPrinters fails, return empty list with error
            logger.error(f'Could not enumerate printers: {str(enum_error)}')
            return JsonResponse({
                'success': False,
                'error': 'Could not enumerate printers. Please check printer configuration.',
                'printers': []
            })
        
        return JsonResponse({
            'success': True,
            'printers': printers,
            'count': len(printers)
        })
        
    except ImportError:
        return JsonResponse({
            'success': False,
            'error': 'Printer detection only works on Windows systems with pywin32 installed. Since your server is on Linux (Digital Ocean), you should run the print client on a Windows PC where printers are connected.',
            'printers': []
        })
    except Exception as e:
        logger.error(f'Error detecting printers: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': 'Error detecting printers. Please try again.',
            'printers': []
        })


@login_required
@require_POST
def regenerate_api_token(request):
    """Regenerate API token for print client authentication"""
    if not (request.user.is_owner() or request.user.is_main_owner() or request.user.is_branch_owner() or request.user.is_manager()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    try:
        from rest_framework.authtoken.models import Token
        
        # Delete old token
        Token.objects.filter(user=request.user).delete()
        
        # Create new token
        token = Token.objects.create(user=request.user)
        
        return JsonResponse({
            'success': True,
            'token': token.key,
            'message': 'API token regenerated successfully. Update your Print Client config.json with the new token.'
        })
    except Exception as e:
        logger.error(f'Error regenerating token: {str(e)}')
        return JsonResponse({
            'success': False,
            'error': 'Error regenerating token. Please try again.'
        }, status=500)
